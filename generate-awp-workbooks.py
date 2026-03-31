#!/usr/bin/env python3
"""Generate IESP AWP Excel workbooks in 14-column working paper format.

v4 — Domain-level conclusions:
- 14 columns: BNM Ref added as column A
- Conclusions at control domain level only (not per sub-item)
- Each BNM sub-item gets its own row as a test procedure under the domain
- Best practice additions have blank BNM Ref
- Scoring Dashboard counts conclusions from domain rows only
- Cloud AWP restructured to match BNM Appendix 10 sub-item structure
- Emerging Tech restructured: 2 BNM domains + 5 best-practice domains
- DCRA, NRA, Digital Services restructured similarly
"""

import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── Styles ──────────────────────────────────────────────────────────
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
SECTION_FONT = Font(name='Calibri', bold=True, size=11, color='2F5496')
SECTION_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
DOMAIN_FONT = Font(name='Calibri', bold=True, size=10)
DOMAIN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
BODY_FONT = Font(name='Calibri', size=10)
CONTEXT_FONT = Font(name='Calibri', size=10, italic=True)
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='2F5496')
SUBTITLE_FONT = Font(name='Calibri', bold=True, size=12, color='2F5496')
LABEL_FONT = Font(name='Calibri', bold=True, size=10)
VALUE_FONT = Font(name='Calibri', size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top')
CENTER_ALIGN = Alignment(horizontal='center', vertical='top', wrap_text=True)

# Scoring fills
COMPLIANT_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
PARTIAL_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
NON_COMPLIANT_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
NA_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
CONCLUSION_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

# 14-column layout
COL_WIDTHS = {
    'A': 14,   # BNM Ref
    'B': 10,   # Ref
    'C': 11,   # Level
    'D': 22,   # Control Domain
    'E': 28,   # Sub-Procedure
    'F': 50,   # Assessment Procedures
    'G': 35,   # Expected Evidence
    'H': 14,   # Method
    'I': 30,   # Procedures Performed
    'J': 25,   # Evidence Obtained
    'K': 12,   # Evidence Ref
    'L': 30,   # Observations
    'M': 16,   # Conclusion
    'N': 30,   # Recommendations
}

ASSESSMENT_HEADERS = [
    'BNM Ref', 'Ref', 'Level', 'Control Domain', 'Sub-Procedure',
    'Assessment Procedures', 'Expected Evidence', 'Method',
    'Procedures Performed', 'Evidence Obtained', 'Evidence Ref',
    'Observations', 'Conclusion', 'Recommendations'
]
NUM_COLS = len(ASSESSMENT_HEADERS)  # 14


def apply_col_widths(ws):
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def write_header_row(ws, row):
    for col, header in enumerate(ASSESSMENT_HEADERS, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def write_section_row(ws, row, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER
    for col in range(2, NUM_COLS + 1):
        ws.cell(row=row, column=col).fill = SECTION_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER


def write_context_row(ws, row, context):
    """Write a merged context row with italic text explaining why/what/risk."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    cell = ws.cell(row=row, column=1, value=context)
    cell.font = CONTEXT_FONT
    cell.alignment = WRAP_ALIGN
    cell.border = THIN_BORDER
    for col in range(2, NUM_COLS + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER


def write_domain_row(ws, row, bnm_ref, ref, level, control_domain, summary):
    """Write a domain header row. Conclusions (col M) and recommendations (col N) are filled here only."""
    data = [bnm_ref, ref, level, control_domain, summary,
            '', '', '', '', '', '', '', '', '']
    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = DOMAIN_FONT
        cell.fill = DOMAIN_FILL
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER
    # Highlight conclusion and recommendation columns for domain rows
    ws.cell(row=row, column=13).fill = CONCLUSION_FILL  # M — Conclusion
    ws.cell(row=row, column=14).fill = CONCLUSION_FILL  # N — Recommendations


def write_subitem_row(ws, row, bnm_ref, ref, sub_proc, procedures, evidence, method='Inspection'):
    """Write a sub-item test procedure row. No conclusion column — observations only."""
    # Columns: A=BNM Ref, B=Ref, C=Level(blank/inherited), D=Control Domain(blank),
    #          E=Sub-Procedure, F=Assessment Procedures, G=Expected Evidence, H=Method,
    #          I-K=auditor fills, L=Observations, M=blank, N=blank
    data = [bnm_ref, ref, '', '', sub_proc, procedures, evidence, method,
            '', '', '', '', '', '']
    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = BODY_FONT
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER


def write_assessment_rows(ws, start_row, sections):
    """Write section headers, context rows, domain rows, and sub-item rows.

    sections format:
    [
        ('Section Title', 'context string', [
            ('DOMAIN', bnm_ref, ref, level, control_domain, summary),
            ('SUB', bnm_ref, ref, sub_proc, procedures, evidence, method),
            ...
        ]),
        ...
    ]
    """
    r = start_row
    for section_data in sections:
        title, context, items = section_data
        write_section_row(ws, r, title)
        r += 1
        if context:
            write_context_row(ws, r, context)
            r += 1
        for item in items:
            row_type = item[0]
            if row_type == 'DOMAIN':
                _, bnm_ref, ref, level, control_domain, summary = item
                write_domain_row(ws, r, bnm_ref, ref, level, control_domain, summary)
            elif row_type == 'SUB':
                _, bnm_ref, ref, sub_proc, procedures, evidence, method = item
                write_subitem_row(ws, r, bnm_ref, ref, sub_proc, procedures, evidence, method)
            r += 1
    return r


# ── Scoping Sheet ───────────────────────────────────────────────────
def add_scoping_sheet(wb, engagement_name, scoping_note=None):
    ws = wb.create_sheet(title='Scoping')
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20

    r = 1
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1, value=f'ENGAGEMENT SCOPING — {engagement_name}').font = TITLE_FONT
    r += 2

    if scoping_note:
        ws.merge_cells(f'A{r}:F{r}')
        cell = ws.cell(row=r, column=1, value=scoping_note)
        cell.font = CONTEXT_FONT
        cell.alignment = WRAP_ALIGN
        r += 2

    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1, value='ASSESSMENT LEVELS').font = SUBTITLE_FONT
    r += 1
    levels = [
        ('ORG', 'Organisation / Enterprise level. Assessed ONCE per engagement. Policies, frameworks, governance structures that apply across all platforms and systems.'),
        ('PLATFORM', 'Platform / CSP level. Assessed PER PLATFORM in scope. If the FI uses AWS and Azure, repeat these test steps for each. Covers platform-specific configurations and controls.'),
        ('WORKLOAD', 'Workload / Application / System level. Assessed PER CRITICAL SYSTEM in scope using sampling. Covers system-specific deployments, configurations, and arrangements.'),
    ]
    for level, desc in levels:
        ws.cell(row=r, column=1, value=level).font = LABEL_FONT
        ws.merge_cells(f'B{r}:F{r}')
        ws.cell(row=r, column=2, value=desc).font = VALUE_FONT
        ws.cell(row=r, column=2).alignment = WRAP_ALIGN
        r += 1
    r += 1

    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1, value='PLATFORMS / CSPs IN SCOPE').font = SUBTITLE_FONT
    r += 1
    headers = ['#', 'Platform / CSP', 'Service Models (IaaS/PaaS/SaaS)', 'Regions', 'Contract Ref', 'Notes']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    r += 1
    for i in range(1, 6):
        for col in range(1, 7):
            c = ws.cell(row=r, column=col, value=str(i) if col == 1 else '')
            c.border = THIN_BORDER
            c.font = VALUE_FONT
        r += 1
    r += 1

    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1, value='CRITICAL SYSTEMS / WORKLOADS IN SCOPE').font = SUBTITLE_FONT
    r += 1
    headers = ['#', 'System / Application Name', 'Platform', 'Criticality', 'Data Classification', 'Notes']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    r += 1
    for i in range(1, 11):
        for col in range(1, 7):
            c = ws.cell(row=r, column=col, value=str(i) if col == 1 else '')
            c.border = THIN_BORDER
            c.font = VALUE_FONT
        r += 1
    r += 1

    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1, value='SAMPLING APPROACH').font = SUBTITLE_FONT
    r += 1
    samples = [
        ('Assessment Period:', '[Start Date] to [End Date]'),
        ('User Access Samples:', '[e.g., 20 user accounts, 10 leavers/transfers]'),
        ('Change Management Samples:', '[e.g., 10 recent changes]'),
        ('Incident Samples:', '[e.g., 15 SOC alerts, all significant incidents in period]'),
        ('Firewall Rule Samples:', '[e.g., 20 rules]'),
        ('Configuration Samples:', '[e.g., 10 network devices]'),
    ]
    for label, val in samples:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.merge_cells(f'B{r}:F{r}')
        ws.cell(row=r, column=2, value=val).font = VALUE_FONT
        r += 1

    return ws


# ── Planning Sheet ──────────────────────────────────────────────────
def add_planning_sheet(wb):
    ws = wb.create_sheet(title='Planning')
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30

    r = 1
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(row=r, column=1, value='PLANNING AND SCOPING CHECKLIST').font = TITLE_FONT
    r += 2

    headers = ['Step', 'Activity', 'Responsible', 'Status', 'Notes']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    r += 1

    planning_steps = [
        ('P-01', 'Confirm engagement trigger and regulatory basis'),
        ('P-02', 'Issue engagement letter with scope, timeline, and deliverables'),
        ('P-03', 'Confirm IESP team qualifications and independence (Part C compliance)'),
        ('P-04', 'Issue Document Request List (DRL) to the FI'),
        ('P-05', 'Obtain and review prior assessment reports and remediation status'),
        ('P-06', 'Identify platforms/CSPs in scope and complete scoping sheet'),
        ('P-07', 'Identify critical systems/workloads in scope'),
        ('P-08', 'Define assessment period and sampling approach'),
        ('P-09', 'Determine engagement mode: design adequacy (pre-impl) or operating effectiveness (attestation)'),
        ('P-10', 'Determine if higher-risk services criteria apply (customer data, cross-border)'),
        ('P-11', 'Schedule site visits / remote assessment sessions'),
        ('P-12', 'Conduct kick-off meeting with FI management'),
        ('P-13', 'Receive and catalogue DRL responses'),
        ('P-14', 'Finalise scoping memorandum'),
    ]
    for ref, activity in planning_steps:
        data = [ref, activity, '', '', '']
        for col, val in enumerate(data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = WRAP_ALIGN
            c.border = THIN_BORDER
        r += 1

    return ws


# ── Reporting Sheet ─────────────────────────────────────────────────
def add_reporting_sheet(wb):
    ws = wb.create_sheet(title='Reporting & Attestation')
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30

    r = 1
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(row=r, column=1, value='REPORTING AND ATTESTATION CHECKLIST').font = TITLE_FONT
    r += 2

    headers = ['Step', 'Activity', 'Responsible', 'Status', 'Notes']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    r += 1

    steps = [
        ('R-01', 'Consolidate all test results — compile conclusions from all domain rows across assessment sheets'),
        ('R-02', 'Assign risk ratings (High/Medium/Low) to each finding based on likelihood and impact'),
        ('R-03', 'Map findings to specific regulatory references (Appendix clauses, Part D items)'),
        ('R-04', 'Draft recommendations with specific remediation steps, responsible parties, and timelines'),
        ('R-05', 'Distinguish must-fix-before-go-live findings from post-launch improvements (for pre-impl)'),
        ('R-06', 'Conduct quality assurance / peer review of working papers and findings'),
        ('R-07', 'Conduct findings walkthrough with FI management'),
        ('R-08', 'Incorporate management responses and agreed action plans'),
        ('R-09', 'Form Part C attestation opinion: Type A (Clean), Type B (With Exceptions), or Type C (Adverse)'),
        ('R-10', 'Prepare the formal report in Appendix 7 Part A format'),
        ('R-11', 'Confirm independence declaration and Part C compliance statement'),
        ('R-12', 'Obtain FI management sign-off on report and management action plans'),
        ('R-13', 'Prepare executive summary for designated board committee (per 8.3)'),
        ('R-14', 'Attend board committee meeting to present findings (if required)'),
        ('R-15', 'Document board committee directives and incorporate into final report'),
        ('R-16', 'Issue final signed report'),
    ]
    for ref, activity in steps:
        data = [ref, activity, '', '', '']
        for col, val in enumerate(data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = WRAP_ALIGN
            c.border = THIN_BORDER
        r += 1

    return ws


# ── Part C Self-Assessment ──────────────────────────────────────────
def add_part_c_sheet(wb):
    ws = wb.create_sheet(title='Part C Self-Assessment')
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30

    r = 1
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1, value='APPENDIX 7 PART C — IESP SELF-ASSESSMENT').font = TITLE_FONT
    r += 1
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1, value='The IESP must comply with these 6 requirements. Complete this self-assessment before issuing the report.').font = VALUE_FONT
    ws.cell(row=r, column=1).alignment = WRAP_ALIGN
    r += 2

    headers = ['#', 'Requirement', 'Assessment Criteria', 'Evidence / Justification', 'Compliant?', 'Notes']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    r += 1

    requirements = [
        ('1', 'Independence', 'The IESP is independent of the FI\'s technology operations and the systems/services being assessed. No conflicts of interest exist between the IESP and the FI.'),
        ('2', 'Competence', 'The IESP team has relevant professional qualifications (CISA, CISSP, CCSP, CSP-specific certs), certifications, and experience in the subject matter being assessed.'),
        ('3', 'Scope', 'The assessment covers ALL areas specified in the relevant appendix (Appendix 10, Appendix 9, or Part D as applicable). No required areas were excluded without documented justification.'),
        ('4', 'Methodology', 'The IESP uses a systematic, risk-based assessment methodology with clear criteria for evaluating controls. The methodology is documented and consistently applied.'),
        ('5', 'Reporting', 'Findings are reported in the Appendix 7 Part A format with clear risk ratings and actionable recommendations. The report is complete, accurate, and traceable to evidence.'),
        ('6', 'Quality Assurance', 'Internal QA processes were followed, including peer review of assessment work by a qualified reviewer who did not perform the assessment.'),
    ]
    for num, req, criteria in requirements:
        data = [num, req, criteria, '', '', '']
        for col, val in enumerate(data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = WRAP_ALIGN
            c.border = THIN_BORDER
        r += 1

    r += 1
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1, value='SIGN-OFF').font = SUBTITLE_FONT
    r += 1
    for label in ['Lead Assessor:', 'Quality Reviewer:']:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value='________________________').font = VALUE_FONT
        ws.cell(row=r, column=4, value='Date:').font = LABEL_FONT
        ws.cell(row=r, column=5, value='______________').font = VALUE_FONT
        r += 1

    return ws


# ── Methodology Sheet ───────────────────────────────────────────────
def create_methodology_sheet(wb, engagement_name, regulatory_basis, scope_desc,
                              assessment_covers, engagement_type_label):
    ws = wb.active
    ws.title = 'Methodology & Approach'
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    for col in 'CDEFGHIJKLMN':
        ws.column_dimensions[col].width = 12

    r = 1
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='INDEPENDENT ASSESSMENT — METHODOLOGY & APPROACH').font = TITLE_FONT
    r += 1
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value=engagement_name).font = SUBTITLE_FONT
    r += 2

    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='ENGAGEMENT DETAILS').font = SUBTITLE_FONT
    r += 1
    details = [
        ('Entity Name:', '[Entity Name]'),
        ('Engagement Type:', engagement_type_label),
        ('Regulatory Basis:', regulatory_basis),
        ('Assessment Date:', '[DD/MM/YYYY]'),
        ('Assessment Period:', '[Start Date] to [End Date]'),
        ('Lead Assessor:', '[Name, Qualification]'),
        ('Engagement Reference:', '[Ref Number]'),
        ('Report Classification:', 'Confidential (Sulit)'),
    ]
    for label, value in details:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=value).font = VALUE_FONT
        r += 1
    r += 1

    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='SCOPE OF ASSESSMENT').font = SUBTITLE_FONT
    r += 1
    for text in [scope_desc, assessment_covers]:
        ws.merge_cells(f'A{r}:N{r}')
        ws.cell(row=r, column=1, value=text).font = VALUE_FONT
        ws.cell(row=r, column=1).alignment = WRAP_ALIGN
        r += 1
    r += 1

    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='ENGAGEMENT MODE').font = SUBTITLE_FONT
    r += 1
    modes = [
        ('Design Adequacy (Pre-Implementation):', 'Assess whether proposed/planned controls are designed to meet the requirement. Evidence is primarily documentary: architecture designs, policies, vendor assessments, staging test results.'),
        ('Operating Effectiveness (Independent Attestation):', 'Assess whether controls operated effectively during the assessment period. Evidence is observation, sampling, re-performance, system-generated logs. Requires defined sampling methodology.'),
    ]
    for label, desc in modes:
        ws.merge_cells(f'A{r}:N{r}')
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        r += 1
        ws.merge_cells(f'A{r}:N{r}')
        ws.cell(row=r, column=1, value=desc).font = VALUE_FONT
        ws.cell(row=r, column=1).alignment = WRAP_ALIGN
        r += 1
    r += 1

    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='ASSESSMENT LEVELS').font = SUBTITLE_FONT
    r += 1
    lvls = [
        ('ORG', 'Organisation-level. Assessed once per engagement (policies, governance, frameworks).'),
        ('PLATFORM', 'Platform/CSP-level. Assessed per platform in scope (configurations, controls per CSP).'),
        ('WORKLOAD', 'System/application-level. Assessed per critical system using sampling.'),
    ]
    for lvl, desc in lvls:
        ws.cell(row=r, column=1, value=lvl).font = LABEL_FONT
        ws.cell(row=r, column=2, value=desc).font = VALUE_FONT
        ws.cell(row=r, column=2).alignment = WRAP_ALIGN
        r += 1
    r += 1

    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='ASSESSMENT METHODS').font = SUBTITLE_FONT
    r += 1
    for method, desc in [
        ('Inspection', 'Examination of documents, records, policies, configurations, and tangible evidence.'),
        ('Inquiry', 'Interviews with management, process owners, and staff.'),
        ('Observation', 'Direct observation of processes, systems, and activities.'),
        ('Confirmation', 'Verification with external or independent parties.'),
        ('Re-performance', 'IESP independently re-performs the control procedure.'),
    ]:
        ws.cell(row=r, column=1, value=method).font = LABEL_FONT
        ws.cell(row=r, column=2, value=desc).font = VALUE_FONT
        ws.cell(row=r, column=2).alignment = WRAP_ALIGN
        r += 1
    r += 1

    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='CONCLUSION SCALE').font = SUBTITLE_FONT
    r += 1
    for conc, desc in [
        ('Compliant', 'Control fully implemented and operating effectively. Evidence supports regulatory criteria are met.'),
        ('Partially Compliant', 'Implemented but with gaps in scope, coverage, documentation, or effectiveness.'),
        ('Non-Compliant', 'Absent, fundamentally deficient, or not operating. Regulatory criteria not met.'),
        ('N/A', 'Not applicable to entity\'s operations. Justification documented.'),
    ]:
        ws.cell(row=r, column=1, value=conc).font = LABEL_FONT
        ws.cell(row=r, column=2, value=desc).font = VALUE_FONT
        ws.cell(row=r, column=2).alignment = WRAP_ALIGN
        r += 1
    r += 1

    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='EVIDENCE HIERARCHY (prefer higher-ranked)').font = SUBTITLE_FONT
    r += 1
    for rank, desc in [
        ('Rank 1', 'Direct observation — IESP directly observes control in operation'),
        ('Rank 2', 'Independent confirmation — Third-party evidence (SOC 2, certifications)'),
        ('Rank 3', 'System-generated — Logs, configurations without manual intervention'),
        ('Rank 4', 'Re-performance — IESP re-performs control procedure'),
        ('Rank 5', 'Documentary — Policies, procedures, meeting minutes'),
        ('Rank 6', 'Inquiry — Verbal representations from FI personnel'),
    ]:
        ws.cell(row=r, column=1, value=rank).font = LABEL_FONT
        ws.cell(row=r, column=2, value=desc).font = VALUE_FONT
        r += 1
    r += 1

    # ── SAMPLING METHODOLOGY ──────────────────────────────────────
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='SAMPLING METHODOLOGY').font = SUBTITLE_FONT
    r += 1

    ws.cell(row=r, column=1, value='Sampling Approach:').font = LABEL_FONT
    ws.merge_cells(f'B{r}:N{r}')
    ws.cell(row=r, column=2, value='Judgmental (not statistical) — appropriate for limited assurance engagements under BNM RMiT. Sample sizes are based on population size and control risk tier.').font = VALUE_FONT
    ws.cell(row=r, column=2).alignment = WRAP_ALIGN
    r += 2

    ws.cell(row=r, column=1, value='Sample Size Table:').font = LABEL_FONT
    r += 1
    sample_headers = ['Population Size', 'Standard Risk', 'High Risk', 'Rationale']
    for col, h in enumerate(sample_headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = LABEL_FONT
        c.border = THIN_BORDER
    r += 1
    sample_rows = [
        ('1-5', 'All', 'All', 'Full coverage feasible'),
        ('6-15', '3', '5', '~25-40% coverage'),
        ('16-50', '5', '8', 'Sufficient for pattern detection'),
        ('51-100', '8', '12', 'Diminishing returns beyond this'),
        ('100+', '10', '15', 'Cap with stratification'),
    ]
    for pop, std, high, rationale in sample_rows:
        data = [pop, std, high, rationale]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = VALUE_FONT
            c.border = THIN_BORDER
            c.alignment = WRAP_ALIGN
        r += 1
    r += 1

    ws.cell(row=r, column=1, value='Time-Based Evidence Coverage:').font = LABEL_FONT
    r += 1
    time_headers = ['Evidence Type', 'Coverage Period', 'Rationale']
    for col, h in enumerate(time_headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = LABEL_FONT
        c.border = THIN_BORDER
    r += 1
    time_rows = [
        ('Board/committee reports', 'Last 4 quarters', 'Full annual governance cycle'),
        ('Operational records (incidents, changes, patches)', 'Last 3-6 months', 'Recent operating effectiveness'),
        ('Annual processes (risk assessment, DR test, pen test)', 'Last 12 months', 'Full cycle'),
        ('Continuous monitoring (SOC, logs, alerts)', 'Last 30 days', 'Current state verification'),
        ('Policies and frameworks', 'Current version + prior', 'Change tracking'),
    ]
    for etype, period, rationale in time_rows:
        data = [etype, period, rationale]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = VALUE_FONT
            c.border = THIN_BORDER
            c.alignment = WRAP_ALIGN
        r += 1
    r += 1

    ws.cell(row=r, column=1, value='Control Risk Tiers:').font = LABEL_FONT
    r += 1
    tier_headers = ['Tier', 'Controls', 'Sampling Level']
    for col, h in enumerate(tier_headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = LABEL_FONT
        c.border = THIN_BORDER
    r += 1
    tier_rows = [
        ('Critical', 'Governance framework, SOC, incident response, regulatory notification, BCM, customer data protection', 'High risk samples'),
        ('Standard', 'Most controls (risk assessment, vuln mgmt, vendor assessment, change mgmt, etc.)', 'Standard risk samples'),
        ('Conditional', 'Emerging tech, specialised areas — may be N/A', 'Standard if applicable'),
    ]
    for tier, controls, sampling in tier_rows:
        data = [tier, controls, sampling]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = VALUE_FONT
            c.border = THIN_BORDER
            c.alignment = WRAP_ALIGN
        r += 1
    r += 1

    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='LIMITATIONS').font = SUBTITLE_FONT
    r += 1
    for lim in [
        '1. Limited assurance engagement — conclusions based on procedures performed, not absolute assurance.',
        '2. Point-in-time (design adequacy) or period assessment (operating effectiveness).',
        '3. Reliance on management representations and documentation provided.',
        '4. Scope guided by BNM RMiT (BNM/RH/PD 028-98, 28 Nov 2025).',
        '5. Does not constitute legal or regulatory advice.',
        '6. The IESP remains solely responsible for the sufficiency of work performed.',
    ]:
        ws.merge_cells(f'A{r}:N{r}')
        ws.cell(row=r, column=1, value=lim).font = VALUE_FONT
        ws.cell(row=r, column=1).alignment = WRAP_ALIGN
        r += 1
    r += 1

    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='SIGN-OFF').font = SUBTITLE_FONT
    r += 1
    for s in ['Lead Assessor:', 'Quality Reviewer:', 'Entity Representative:']:
        ws.cell(row=r, column=1, value=s).font = LABEL_FONT
        ws.cell(row=r, column=2, value='________________________').font = VALUE_FONT
        ws.cell(row=r, column=6, value='Date:').font = LABEL_FONT
        ws.cell(row=r, column=7, value='______________').font = VALUE_FONT
        r += 1
    return ws


# ── Part D (shared) ─────────────────────────────────────────────────
PART_D_SECTIONS = [
    ('Item 1(a) — Access Control',
     'Access control is foundational — who can access what. Expect policy-driven provisioning, periodic reviews, PAM controls, timely revocation. Excessive access is the most common audit finding.',
     [
        ('DOMAIN', 'App7-D1(a)', 'PD-01', 'ORG', 'Access Control', 'Access control policies, provisioning, privileged access, and periodic review'),
        ('SUB', 'App7-D1(a)(i)', 'PD-01.1', 'Access control policies and mechanisms',
         '1. Obtain the FI\'s access control policy and supporting standards.\n2. Verify policy covers user registration, de-registration, access provisioning, and privilege management.\n3. Sample 20 user accounts — verify access rights are consistent with job roles (least privilege).\n4. Check that access reviews are performed periodically (quarterly for privileged, semi-annually for standard).\n5. Verify access revocation upon termination — sample 10 recent leavers/transfers.',
         'Access control policy; User access provisioning records; Access review reports; HR leaver/transfer list vs. access revocation logs', 'Inspection'),
        ('SUB', 'App7-D1(a)(ii)', 'PD-01.2', 'Privileged access management',
         '1. Obtain the PAM policy or standard.\n2. Obtain the list of all privileged accounts (system admin, DBA, root, domain admin).\n3. Verify privileged accounts are individually assigned or vault-managed with session recording.\n4. Check that privileged access is time-bound and subject to approval workflow.\n5. Review privileged session logs for the past 3 months for anomalies.',
         'PAM policy; Privileged account inventory; PAM tool configuration and logs; Session recording samples', 'Inspection'),
     ]),
    ('Item 1(b) — Physical and Environmental Security',
     'Physical controls protect technology assets from unauthorised access and environmental threats. Expect multi-layered access, CCTV, environmental monitoring. Physical compromise bypasses logical controls.',
     [
        ('DOMAIN', 'App7-D1(b)', 'PD-02', 'ORG', 'Physical Security', 'Physical access controls and environmental protection for technology assets'),
        ('SUB', 'App7-D1(b)(i)', 'PD-02.1', 'Physical security controls',
         '1. Identify all sensitive areas (data centres, server rooms, network rooms).\n2. Verify multi-layered physical access controls (badge, biometric, PIN).\n3. Review visitor management procedures.\n4. Verify CCTV coverage with minimum 90-day retention.\n5. Sample-check physical access logs against approved access lists.',
         'Physical security policy; Access control system config; Visitor logs; CCTV coverage map; Physical access log reconciliation', 'Inspection / Observation'),
        ('SUB', 'App7-D1(b)(ii)', 'PD-02.2', 'Environmental controls',
         '1. Verify fire detection and suppression systems are current on inspections.\n2. Check water leak detection.\n3. Verify temperature and humidity monitoring with alerts.\n4. Confirm environmental incident response procedures are tested.',
         'Fire system inspection certificates; Water leak detection layout; Environmental monitoring dashboards; Incident response test records', 'Inspection / Observation'),
     ]),
    ('Item 1(c) — Operations Security',
     'Operational controls ensure systems are managed securely. Expect documented procedures, controlled changes, tested backups, malware protection, vulnerability management. Operational failures cause avoidable incidents.',
     [
        ('DOMAIN', 'App7-D1(c)', 'PD-03', 'ORG', 'Operations Security', 'Operational procedures, change management, backup, malware protection, and vulnerability management'),
        ('SUB', 'App7-D1(c)(i)', 'PD-03.1', 'Operational procedures and controls',
         '1. Obtain documented operating procedures (change mgmt, capacity mgmt, backup, malware protection).\n2. Verify change management — sample 10 changes with full lifecycle.\n3. Verify backup procedures — confirm critical systems backed up and restore tested.\n4. Verify anti-malware deployed with current signatures.',
         'Operating procedures; 10 sampled change records; Backup schedule and restore test results; Anti-malware deployment reports', 'Inspection'),
        ('SUB', 'App7-D1(c)(ii)', 'PD-03.2', 'Vulnerability management and patching',
         '1. Obtain vulnerability management policy and patching SLAs.\n2. Review most recent vulnerability scan results.\n3. Check patch compliance rates.\n4. Verify exceptions are formally risk-accepted.',
         'Vulnerability management policy; Scan reports; Patch compliance dashboard; Risk acceptance records', 'Inspection'),
     ]),
    ('Item 1(d) — Communications Security',
     'Network controls protect data in transit and network integrity. Expect segmentation, least-privilege firewall rules, encryption, IDS/IPS. Network compromise provides access to all connected systems.',
     [
        ('DOMAIN', 'App7-D1(d)', 'PD-04', 'ORG', 'Communications Security', 'Network security, segmentation, encryption of data in transit, and intrusion detection'),
        ('SUB', 'App7-D1(d)(i)', 'PD-04.1', 'Network security and secure communications',
         '1. Review network segmentation.\n2. Verify firewall rules — sample 20 rules for least-privilege.\n3. Confirm encryption of data in transit (TLS 1.2+).\n4. Verify secure configuration of email, messaging, file transfer.\n5. Check IDS/IPS deployed at critical boundaries.',
         'Network architecture diagram; Firewall rule base sample; TLS configuration; IDS/IPS deployment and alert records', 'Inspection'),
     ]),
    ('Item 1(e) — Incident Management',
     'Incident management ensures timely detection, response, and recovery. Expect 24/7 SOC, SIEM, IR playbooks, annual exercises. Without incident management, breaches escalate unchecked.',
     [
        ('DOMAIN', 'App7-D1(e)', 'PD-05', 'ORG', 'Incident Management', 'Incident detection, classification, response, RCA, and regulatory notification'),
        ('SUB', 'App7-D1(e)(i)', 'PD-05.1', 'Incident management framework',
         '1. Obtain incident management policy and procedures.\n2. Verify incident classification scheme.\n3. Review incident log for past 12 months.\n4. Verify RCA for significant incidents with corrective actions.\n5. Confirm regulatory notification procedures.',
         'Incident management policy; Classification matrix; Incident log (12 months); RCA reports; Regulatory notification records', 'Inspection'),
        ('SUB', 'App7-D1(e)(ii)', 'PD-05.2', 'Detection and response capabilities',
         '1. Confirm SOC operates 24/7.\n2. Verify SIEM deployment with critical log sources.\n3. Sample 15 alerts — verify triage.\n4. Confirm IR playbooks for common scenarios.\n5. Verify annual IR drills/tabletop exercises.',
         'SOC operational documentation; SIEM log source inventory; 15 sampled alert records; IR playbooks; Drill/tabletop reports', 'Inspection / Inquiry'),
     ]),
    ('Item 1(f) — Business Continuity',
     'Business continuity ensures operations continue during disruptions. Expect BCP/DRP with cyber scenarios, tested DR including security controls, redundant security infrastructure. Untested DR fails when needed.',
     [
        ('DOMAIN', 'App7-D1(f)', 'PD-06', 'ORG', 'Business Continuity', 'Security in BCP/DRP, DR testing with security controls, and redundancy of security infrastructure'),
        ('SUB', 'App7-D1(f)(i)', 'PD-06.1', 'Security in BCP/DRP',
         '1. Obtain BCP and IT DRP.\n2. Verify security requirements addressed in BCP/DRP.\n3. Check BCP/DRP includes cyber incident scenarios.\n4. Verify recovery procedures include security validation.',
         'BCP and IT DRP; Security requirements in BCP/DRP; Cyber incident scenarios; Recovery security validation checklists', 'Inspection'),
        ('SUB', 'App7-D1(f)(ii)', 'PD-06.2', 'BCP/DRP testing with security controls',
         '1. Review most recent BCP/DRP test.\n2. Verify security controls tested during DR exercise.\n3. Confirm security findings remediated.\n4. Verify DR site maintains equivalent security posture.',
         'BCP/DRP test plan and results; Security test components; DR site security assessment; Remediation tracker', 'Inspection'),
        ('SUB', 'App7-D1(f)(iii)', 'PD-06.3', 'Redundancy of security infrastructure',
         '1. Identify critical security infrastructure and verify HA configuration.\n2. Confirm failover tested.\n3. Verify monitoring continues during DR.\n4. Check licenses valid at DR site.',
         'Security infrastructure HA architecture; Failover test records; DR monitoring evidence; License/subscription records', 'Inspection'),
     ]),
    ('Item 2(a) — Customer Identity Authentication',
     'Customer authentication protects online services from unauthorised access. Expect MFA, secure sessions, strong cryptography, anti-session-fixation. Weak authentication directly exposes customers.',
     [
        ('DOMAIN', 'App7-D2(a)', 'PD-07', 'WORKLOAD', 'Customer Authentication', 'Session protection, MFA, credential storage, and cryptographic implementation'),
        ('SUB', 'App7-D2(a)(i)', 'PD-07.1', 'Session and MITM protection',
         '1. Review session management controls.\n2. Verify TLS 1.2+, HSTS, certificate validity.\n3. Confirm session tokens are cryptographically random with proper expiry.\n4. Check anti-session-fixation controls.\n5. Verify certificate pinning for mobile apps.',
         'TLS scan results; Session management config; Application security assessment; Mobile cert pinning evidence', 'Inspection'),
        ('SUB', 'App7-D2(a)(ii)', 'PD-07.2', 'Internal controls for online systems',
         '1. Verify application hardening.\n2. Review app security testing results (SAST, DAST, pen test).\n3. Confirm database security controls.\n4. Check segmentation from corporate network.',
         'Hardening standards; App security test reports; Database security config; Network segmentation evidence', 'Inspection'),
        ('SUB', 'App7-D2(a)(iii)', 'PD-07.3', 'Multi-level authentication',
         '1. Review authentication architecture.\n2. Verify MFA for login and high-risk transactions.\n3. Check out-of-band authentication mechanisms.\n4. Verify real-time verification for high-value transactions.\n5. Assess authentication strength relative to transaction risk.',
         'Authentication architecture; MFA config; OTP/out-of-band mechanism; Risk-based authentication rules', 'Inspection'),
        ('SUB', 'App7-D2(a)(iv)', 'PD-07.4', 'Session handling and credential storage',
         '1. Verify session timeouts (idle and absolute).\n2. Confirm concurrent session controls.\n3. Verify credentials stored with strong hashing (bcrypt, Argon2).\n4. Check auth database access-restricted and encrypted.\n5. Review account lockout configuration.',
         'Session management config; Password hashing evidence; Auth database access controls; Account lockout config', 'Inspection'),
        ('SUB', 'App7-D2(a)(v)', 'PD-07.5', 'Cryptographic implementation',
         '1. Review cryptographic standards (AES-256, RSA-2048+, ECDSA P-256+).\n2. Verify key management practices.\n3. Check no deprecated algorithms.\n4. Verify crypto implementation reviewed/tested.',
         'Cryptographic standards document; Key management procedures; HSM config; Crypto assessment results', 'Inspection'),
     ]),
    ('Item 2(b) — Transaction Authentication',
     'Transaction authentication ensures non-repudiation and integrity. Expect transaction signing, tamper-evident audit trails, mutual authentication. Without non-repudiation, disputes cannot be resolved.',
     [
        ('DOMAIN', 'App7-D2(b)', 'PD-08', 'WORKLOAD', 'Transaction Authentication', 'Proof of origin, integrity, secure delivery, and mutual authentication'),
        ('SUB', 'App7-D2(b)(i)', 'PD-08.1', 'Proof of origin and integrity',
         '1. Review transaction signing/MAC mechanisms.\n2. Verify records include origin, timestamp, hash, integrity verification.\n3. Confirm tamper-evident audit trails.\n4. Check logs sufficient for dispute resolution.',
         'Transaction authentication mechanism; Sample transaction records; Audit trail controls; Log retention policy', 'Inspection'),
        ('SUB', 'App7-D2(b)(ii)', 'PD-08.2', 'Secure delivery and mutual authentication',
         '1. Verify delivery channel secured end-to-end.\n2. Check alerts for high-risk transaction types.\n3. Verify mutual authentication for B2B channels.\n4. Review triggers for additional authentication.',
         'Channel security architecture; Transaction alert config; Mutual authentication implementation; Authentication trigger criteria', 'Inspection'),
        ('SUB', 'App7-D2(b)(iii)', 'PD-08.3', 'Transaction notification and confirmation',
         '1. Verify transaction notifications sent (push, SMS, email).\n2. Confirm high-risk transactions trigger real-time alerts.\n3. Check customer can dispute/report through notifications.\n4. Verify notification delivery is reliable.',
         'Transaction notification config; Alert rules; Customer dispute channel; Notification delivery SLAs', 'Inspection'),
        ('SUB', 'App7-D2(b)(iv)', 'PD-08.4', 'Transaction binding and anti-replay',
         '1. Verify authentication tokens bound to specific transaction parameters.\n2. Check anti-replay mechanisms (nonce, timestamp).\n3. Verify OTP/TAC is single-use and expires within 2-5 minutes.\n4. Confirm transaction cannot be modified after authentication.',
         'Transaction binding mechanism; Anti-replay controls; OTP/TAC config; Transaction immutability controls', 'Inspection'),
     ]),
    ('Item 2(c) — Segregation of Duties',
     'SoD prevents single-person control over critical operations. Expect maker-checker, privileged user reviews, tamper-resistant authorisation. SoD failures enable fraud.',
     [
        ('DOMAIN', 'App7-D2(c)', 'PD-09', 'WORKLOAD', 'Segregation of Duties', 'Dual control, maker-checker, and authorisation database integrity'),
        ('SUB', 'App7-D2(c)(i)', 'PD-09.1', 'Dual control for online transactions',
         '1. Identify critical functions in online transaction systems.\n2. Verify maker-checker for critical operations.\n3. Review access control matrices — no single user can initiate and approve.\n4. Verify unauthorised access detection.',
         'Access control matrices; Maker-checker config; Dual control workflow; Unauthorised access detection controls', 'Inspection'),
        ('SUB', 'App7-D2(c)(ii)', 'PD-09.2', 'Authorisation database integrity',
         '1. Verify authorisation database is tamper-resistant.\n2. Confirm changes require approval and are logged.\n3. Obtain privileged user list — verify quarterly review.\n4. Check reviews result in revocation of unnecessary access.',
         'Auth database access controls; Change logs; Privileged user review records; Revocation evidence', 'Inspection'),
        ('SUB', 'App7-D2(c)(iii)', 'PD-09.3', 'SoD in system administration',
         '1. Verify system administration duties separated (DBA vs network vs security).\n2. Check no single admin has unrestricted access across domains.\n3. Verify SoD matrix enforced in IAM.\n4. Review compensating controls where SoD cannot be fully achieved.',
         'SoD matrix; Admin role definitions; IAM enforcement evidence; Compensating controls documentation', 'Inspection'),
        ('SUB', 'App7-D2(c)(iv)', 'PD-09.4', 'Monitoring of privileged activities',
         '1. Verify all privileged activities are logged.\n2. Check logs reviewed regularly (at least weekly).\n3. Confirm anomalous privileged activities trigger alerts.\n4. Verify log integrity (tamper-proof).',
         'Privileged activity logs; Log review schedule; Alerting rules; Log integrity controls', 'Inspection'),
     ]),
    ('Item 2(d) — Data Integrity',
     'Data integrity controls prevent tampering and ensure confidentiality. Expect E2E encryption, defence-in-depth, pen testing, audit trails. Data integrity failure causes financial loss and regulatory breach.',
     [
        ('DOMAIN', 'App7-D2(d)', 'PD-10', 'ORG', 'Data Integrity', 'End-to-end encryption, multi-layer security, SPOF avoidance, security assessment, and audit trails'),
        ('SUB', 'App7-D2(d)(i)', 'PD-10.1', 'End-to-end encryption and multi-layer security',
         '1. Verify E2E encryption for external communications.\n2. Review multi-layer network security (defence-in-depth).\n3. Confirm each layer independently managed.\n4. Verify security devices current on firmware/signatures.',
         'Encryption architecture; Multi-layer security diagram; Security device inventory; Management and monitoring evidence', 'Inspection'),
        ('SUB', 'App7-D2(d)(ii)', 'PD-10.2', 'Absence of single points of failure',
         '1. Review network architecture for SPOF.\n2. Verify redundancy at each critical layer.\n3. Confirm failover tested within 12 months.\n4. Check SPOF analysis documented.',
         'Network architecture with redundancy; SPOF analysis; Failover test results; SPOF review history', 'Inspection'),
        ('SUB', 'App7-D2(d)(iii)', 'PD-10.3', 'Security assessment and audit trails',
         '1. Obtain most recent pen test and network security assessment.\n2. Verify testing by qualified party within 12 months.\n3. Confirm vulnerabilities remediated or risk-accepted.\n4. Verify audit trail capabilities.\n5. Confirm logs tamper-protected and retained.',
         'Pen test report; Network security assessment; Audit trail config and sample logs; Log integrity controls', 'Inspection'),
        ('SUB', 'App7-D2(d)(iv)', 'PD-10.4', 'Data confidentiality and risk-based auth',
         '1. Verify confidentiality controls (encryption, masking, access controls).\n2. Confirm data classification applied.\n3. Verify risk-based step-up authentication.\n4. Check customers receive timely transaction notifications.',
         'Data classification scheme; Encryption and masking config; Risk-based auth rules; Customer notification samples', 'Inspection'),
        ('SUB', 'App7-D2(d)(v)', 'PD-10.5', 'Data at rest protection',
         '1. Verify encryption at rest for databases, file systems, backups.\n2. Check key management (rotation, access controls, HSM).\n3. Verify data masking in non-production environments.\n4. Confirm data disposal procedures.',
         'Encryption at rest config; Key management procedures; Data masking evidence; Data disposal procedures', 'Inspection'),
        ('SUB', 'App7-D2(d)(vi)', 'PD-10.6', 'Integrity monitoring and checksums',
         '1. Verify file integrity monitoring for critical system files.\n2. Check database integrity controls (checksums, referential integrity).\n3. Verify integrity checks in data transfer processes.\n4. Confirm integrity violations trigger alerts.',
         'FIM config; Database integrity controls; Data transfer integrity checks; Integrity alert config', 'Inspection'),
        ('SUB', 'App7-D2(d)(vii)', 'PD-10.7', 'Data loss prevention',
         '1. Verify DLP controls for sensitive data.\n2. Check DLP coverage (email, web, endpoint, cloud).\n3. Review DLP policy rules and exceptions.\n4. Sample DLP incidents (5 recent) — verify investigation.',
         'DLP policy and config; DLP coverage matrix; DLP incident samples; Investigation records', 'Inspection'),
        ('SUB', 'App7-D2(d)(viii)', 'PD-10.8', 'Backup integrity verification',
         '1. Verify backup integrity checks (checksum verification).\n2. Confirm restore testing includes data integrity validation.\n3. Check backup encryption and access controls.\n4. Verify backup logs and monitoring.',
         'Backup integrity checks; Restore test with validation; Backup encryption config; Backup monitoring', 'Inspection'),
     ]),
    ('Item 2(e) — Mobile Device Risks',
     'Mobile apps require additional protections due to the untrusted device environment. Expect jailbreak detection, secure storage, certificate pinning, fake app monitoring. Mobile-specific attacks bypass server-side controls.',
     [
        ('DOMAIN', 'App7-D2(e)', 'PD-11', 'WORKLOAD', 'Mobile Device Security', 'Mobile app security baseline, data protection, transaction controls, and app distribution security'),
        ('SUB', 'App7-D2(e)(i)', 'PD-11.1', 'Mobile app security baseline',
         '1. Confirm minimum OS version enforced and jailbreak/root detection.\n2. Review most recent mobile pen test.\n3. Verify vulnerabilities remediated.\n4. Confirm secure E2E communication (cert pinning, TLS 1.2+).',
         'OS version enforcement; Jailbreak/root detection; Mobile pen test report; Remediation evidence; Communication security config', 'Inspection'),
        ('SUB', 'App7-D2(e)(ii)', 'PD-11.2', 'Mobile data protection',
         '1. Verify sensitive data not stored on device or only in secure enclaves.\n2. Confirm cache/logs sanitised.\n3. Verify clipboard access restricted.\n4. Check remote wipe capability.',
         'Mobile app security architecture; Data storage review; Secure enclave usage; Remote wipe documentation', 'Inspection'),
        ('SUB', 'App7-D2(e)(iii)', 'PD-11.3', 'Transaction notification and codes',
         '1. Verify transaction notifications (push, SMS, email).\n2. Verify suspicious transaction alerts.\n3. Confirm unique per-transaction codes (TAC, OTP).\n4. Verify code expiry (2-5 min) and single-use.',
         'Transaction notification config; Suspicious alert rules; Code generation mechanism; Code expiry config', 'Inspection'),
        ('SUB', 'App7-D2(e)(iv)', 'PD-11.4', 'App distribution and fake app monitoring',
         '1. Verify publishing access controls.\n2. Confirm monitoring for fake/fraudulent apps.\n3. Verify takedown process exists and exercised.\n4. Review fake app incident history.',
         'Publishing access controls; Fake app monitoring service; Takedown process docs; Fake app incident log', 'Inspection'),
        ('SUB', 'App7-D2(e)(v)', 'PD-11.5', 'Code obfuscation and anti-tampering',
         '1. Verify code obfuscation applied to mobile app binaries.\n2. Check anti-tampering and anti-debugging protections.\n3. Verify integrity checks on app launch.\n4. Confirm RASP deployment.',
         'Obfuscation evidence; Anti-tampering config; Integrity check mechanism; RASP deployment', 'Inspection'),
        ('SUB', 'App7-D2(e)(vi)', 'PD-11.6', 'Secure SDK and third-party library management',
         '1. Inventory third-party SDKs and libraries in mobile app.\n2. Verify SDKs from trusted sources with known vulnerabilities checked.\n3. Check data sharing by SDKs is controlled.\n4. Verify library update process.',
         'SDK/library inventory; Vulnerability assessment; Data sharing controls; Library update procedures', 'Inspection'),
        ('SUB', 'App7-D2(e)(vii)', 'PD-11.7', 'Biometric authentication on mobile',
         '1. Verify biometric authentication uses platform secure enclave.\n2. Check fallback authentication mechanism.\n3. Verify biometric data not stored by the app.\n4. Confirm biometric enrollment verification.',
         'Biometric implementation; Secure enclave usage; Fallback mechanism; Enrollment verification process', 'Inspection'),
        ('SUB', 'App7-D2(e)(viii)', 'PD-11.8', 'Push notification security',
         '1. Verify push notifications do not contain sensitive data.\n2. Check notification delivery uses secure channels.\n3. Verify notification content is generic until authenticated.\n4. Confirm user opt-out capability.',
         'Push notification policy; Notification content review; Delivery channel security; Opt-out mechanism', 'Inspection'),
        ('SUB', 'App7-D2(e)(ix)', 'PD-11.9', 'Device binding and registration',
         '1. Verify device binding mechanism.\n2. Check device registration process requires authentication.\n3. Verify device change triggers step-up verification.\n4. Confirm maximum device limits enforced.',
         'Device binding mechanism; Registration flow; Device change alerts; Device limit config', 'Inspection'),
        ('SUB', 'App7-D2(e)(x)', 'PD-11.10', 'Secure communication channel',
         '1. Verify certificate pinning implemented.\n2. Check for SSL/TLS downgrade protection.\n3. Verify proxy detection.\n4. Confirm mutual TLS where applicable.',
         'Certificate pinning config; Downgrade protection; Proxy detection mechanism; mTLS config', 'Inspection'),
        ('SUB', 'App7-D2(e)(xi)', 'PD-11.11', 'Mobile app update and versioning',
         '1. Verify forced update mechanism for critical security patches.\n2. Check minimum supported app version policy.\n3. Verify graceful degradation for unsupported versions.\n4. Confirm update notification to users.',
         'Forced update mechanism; Minimum version policy; Degradation handling; Update notification config', 'Inspection'),
     ]),
]


def add_part_d_sheet(wb, sheet_name='Appendix 7 Part D'):
    ws = wb.create_sheet(title=sheet_name)
    apply_col_widths(ws)
    r = 1
    write_header_row(ws, r)
    r += 1
    write_assessment_rows(ws, r, PART_D_SECTIONS)
    return ws


# ── Scoring Dashboard ──────────────────────────────────────────────
NON_ASSESSMENT_SHEETS = {
    'Methodology & Approach', 'Scoping', 'Planning',
    'Reporting & Attestation', 'Part C Self-Assessment', 'Scoring Dashboard',
}


def add_scoring_sheet(wb):
    """Add a Scoring Dashboard sheet with per-sheet summaries, overall totals,
    Part C opinion indicator, and level breakdown.

    Conclusions are counted from DOMAIN rows only — identified by Ref (col B)
    matching pattern XX-NN (no dot), and Conclusion is in col M."""
    ws = wb.create_sheet(title='Scoring Dashboard')
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 18

    assessment_sheets = [
        s for s in wb.sheetnames if s not in NON_ASSESSMENT_SHEETS
    ]

    r = 1
    # ── Title ──
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1, value='SCORING DASHBOARD').font = TITLE_FONT
    r += 2

    # ── Explanation ──
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1,
            value='Conclusions are at the DOMAIN level only (Ref pattern: XX-NN without a dot). '
                  'Sub-item rows (XX-NN.N) have observations but no conclusion.').font = CONTEXT_FONT
    ws.cell(row=r, column=1).alignment = WRAP_ALIGN
    r += 2

    # ── Overall Summary ──
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1, value='OVERALL SUMMARY').font = SUBTITLE_FONT
    r += 1

    overall_headers = ['Metric', 'Value']
    for col, h in enumerate(overall_headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    r += 1

    overall_start_row = r

    overall_labels = [
        'Total Compliant',
        'Total Partially Compliant',
        'Total Non-Compliant',
        'Total N/A',
        'Total Not Assessed',
        'Total Control Domains',
        'Overall Compliance %',
    ]
    for label in overall_labels:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2).font = VALUE_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        r += 1

    r += 1
    # Part C opinion indicator
    ws.merge_cells(f'A{r}:B{r}')
    ws.cell(row=r, column=1, value='PART C OPINION INDICATOR').font = SUBTITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value='Opinion Type').font = LABEL_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2).border = THIN_BORDER
    opinion_row = r
    r += 1

    ws.cell(row=r, column=1, value='Criteria:').font = CONTEXT_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.merge_cells(f'B{r}:H{r}')
    ws.cell(row=r, column=2,
            value='Type A (Clean): NC=0 and PC<=3  |  '
                  'Type B (With Exceptions): NC>0 and NC<=5  |  '
                  'Type C (Adverse): NC>5').font = CONTEXT_FONT
    ws.cell(row=r, column=2).alignment = WRAP_ALIGN
    ws.cell(row=r, column=2).border = THIN_BORDER
    r += 2

    # ── Per-Sheet Summary ──
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1, value='PER-SHEET BREAKDOWN').font = SUBTITLE_FONT
    r += 1

    ps_headers = ['Sheet', 'Compliant', 'Partially Compliant',
                  'Non-Compliant', 'N/A', 'Not Assessed', 'Total Domains',
                  'Compliance %']
    for col, h in enumerate(ps_headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = CENTER_ALIGN
    r += 1

    per_sheet_first_row = r

    for sheet_name in assessment_sheets:
        ws.cell(row=r, column=1, value=sheet_name).font = LABEL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER

        safe = sheet_name.replace("'", "''")

        # Domain rows have Ref in col B matching pattern without dots.
        # Conclusions are in col M for domain rows.
        # We use COUNTIFS: col B matches pattern (non-blank, no dot) AND col M matches conclusion.
        # Since COUNTIFS can't do "not contains dot", we count conclusions from col M
        # where col B is non-blank. Domain rows are the only ones with col M filled.
        # The template sets up domain rows with highlighted M column.
        # Simply COUNTIF on col M works because only domain rows have conclusions.

        # Compliant count — COUNTIF on column M
        ws.cell(row=r, column=2).font = VALUE_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=2).fill = COMPLIANT_FILL
        ws.cell(row=r, column=2,
                value=f"=COUNTIF('{safe}'!M:M,\"Compliant\")")

        # Partially Compliant
        ws.cell(row=r, column=3).font = VALUE_FONT
        ws.cell(row=r, column=3).border = THIN_BORDER
        ws.cell(row=r, column=3).fill = PARTIAL_FILL
        ws.cell(row=r, column=3,
                value=f"=COUNTIF('{safe}'!M:M,\"Partially Compliant\")")

        # Non-Compliant
        ws.cell(row=r, column=4).font = VALUE_FONT
        ws.cell(row=r, column=4).border = THIN_BORDER
        ws.cell(row=r, column=4).fill = NON_COMPLIANT_FILL
        ws.cell(row=r, column=4,
                value=f"=COUNTIF('{safe}'!M:M,\"Non-Compliant\")")

        # N/A
        ws.cell(row=r, column=5).font = VALUE_FONT
        ws.cell(row=r, column=5).border = THIN_BORDER
        ws.cell(row=r, column=5).fill = NA_FILL
        ws.cell(row=r, column=5,
                value=f"=COUNTIF('{safe}'!M:M,\"N/A\")")

        # Not Assessed: domain rows with non-blank B and blank M
        # Domain rows have B non-blank AND C non-blank (level is set on domain rows)
        # Sub-item rows have C blank. So COUNTIFS(B<>"", C<>"", M="") counts domain rows not assessed.
        ws.cell(row=r, column=6).font = VALUE_FONT
        ws.cell(row=r, column=6).border = THIN_BORDER
        ws.cell(row=r, column=6,
                value=f"=COUNTIFS('{safe}'!B:B,\"<>\",'{safe}'!C:C,\"<>\",'{safe}'!M:M,\"=\")"
                      f"-COUNTIF('{safe}'!B:B,\"Ref\")")

        # Total domains (rows with non-blank B and non-blank C, minus header)
        ws.cell(row=r, column=7).font = VALUE_FONT
        ws.cell(row=r, column=7).border = THIN_BORDER
        ws.cell(row=r, column=7,
                value=f"=COUNTIFS('{safe}'!B:B,\"<>\",'{safe}'!C:C,\"<>\")"
                      f"-COUNTIF('{safe}'!B:B,\"Ref\")")

        # Compliance % = Compliant / (Total - N/A - Not Assessed)
        b_col = f'B{r}'
        g_col = f'G{r}'
        e_col = f'E{r}'
        f_col = f'F{r}'
        ws.cell(row=r, column=8).font = VALUE_FONT
        ws.cell(row=r, column=8).border = THIN_BORDER
        ws.cell(row=r, column=8).number_format = '0.0%'
        ws.cell(row=r, column=8,
                value=f'=IF(({g_col}-{e_col}-{f_col})=0,"\u2014",'
                      f'{b_col}/({g_col}-{e_col}-{f_col}))')

        r += 1

    per_sheet_last_row = r - 1

    # ── Fill Overall Summary formulas ──
    ps_range_b = f'B{per_sheet_first_row}:B{per_sheet_last_row}'
    ps_range_c = f'C{per_sheet_first_row}:C{per_sheet_last_row}'
    ps_range_d = f'D{per_sheet_first_row}:D{per_sheet_last_row}'
    ps_range_e = f'E{per_sheet_first_row}:E{per_sheet_last_row}'
    ps_range_f = f'F{per_sheet_first_row}:F{per_sheet_last_row}'
    ps_range_g = f'G{per_sheet_first_row}:G{per_sheet_last_row}'

    or_row = overall_start_row  # Total Compliant
    ws.cell(row=or_row, column=2, value=f'=SUM({ps_range_b})')
    ws.cell(row=or_row, column=2).fill = COMPLIANT_FILL
    or_row += 1  # Total Partially Compliant
    ws.cell(row=or_row, column=2, value=f'=SUM({ps_range_c})')
    ws.cell(row=or_row, column=2).fill = PARTIAL_FILL
    or_row += 1  # Total Non-Compliant
    ws.cell(row=or_row, column=2, value=f'=SUM({ps_range_d})')
    ws.cell(row=or_row, column=2).fill = NON_COMPLIANT_FILL
    nc_cell = f'B{or_row}'
    or_row += 1  # Total N/A
    ws.cell(row=or_row, column=2, value=f'=SUM({ps_range_e})')
    ws.cell(row=or_row, column=2).fill = NA_FILL
    na_total_cell = f'B{or_row}'
    or_row += 1  # Total Not Assessed
    ws.cell(row=or_row, column=2, value=f'=SUM({ps_range_f})')
    not_assessed_cell = f'B{or_row}'
    or_row += 1  # Total Control Domains
    ws.cell(row=or_row, column=2, value=f'=SUM({ps_range_g})')
    total_cell = f'B{or_row}'
    or_row += 1  # Overall Compliance %
    compliant_cell = f'B{overall_start_row}'
    ws.cell(row=or_row, column=2).number_format = '0.0%'
    ws.cell(row=or_row, column=2,
            value=f'=IF(({total_cell}-{na_total_cell}-{not_assessed_cell})=0,"\u2014",'
                  f'{compliant_cell}/({total_cell}-{na_total_cell}-{not_assessed_cell}))')

    # Part C opinion IF formula
    pc_cell = f'B{overall_start_row + 1}'  # Partially Compliant total
    ws.cell(row=opinion_row, column=2,
            value=f'=IF({nc_cell}>5,"Type C (Adverse)",'
                  f'IF({nc_cell}>0,"Type B (With Exceptions)",'
                  f'IF({pc_cell}<=3,"Type A (Clean)","Type B (With Exceptions)")))')

    r += 2

    # ── Level Breakdown ──
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1, value='LEVEL BREAKDOWN').font = SUBTITLE_FONT
    r += 1

    lvl_headers = ['Level', 'Compliant', 'Partially Compliant',
                   'Non-Compliant', 'N/A', 'Not Assessed', 'Total', '']
    for col, h in enumerate(lvl_headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = CENTER_ALIGN
    r += 1

    for level in ['ORG', 'PLATFORM', 'WORKLOAD']:
        ws.cell(row=r, column=1, value=level).font = LABEL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER

        # Level is in col C, conclusion in col M (domain rows only)
        conclusions = [
            ('Compliant', 2, COMPLIANT_FILL),
            ('Partially Compliant', 3, PARTIAL_FILL),
            ('Non-Compliant', 4, NON_COMPLIANT_FILL),
            ('N/A', 5, NA_FILL),
        ]
        for conclusion, col_idx, fill in conclusions:
            parts = []
            for sn in assessment_sheets:
                safe = sn.replace("'", "''")
                parts.append(
                    f"COUNTIFS('{safe}'!C:C,\"{level}\",'{safe}'!M:M,\"{conclusion}\")"
                )
            formula = '=' + '+'.join(parts) if parts else '=0'
            c = ws.cell(row=r, column=col_idx, value=formula)
            c.font = VALUE_FONT
            c.border = THIN_BORDER
            c.fill = fill

        # Not Assessed by level: rows with matching level in C and blank in M
        na_parts = []
        for sn in assessment_sheets:
            safe = sn.replace("'", "''")
            na_parts.append(
                f"COUNTIFS('{safe}'!C:C,\"{level}\",'{safe}'!M:M,\"=\")"
            )
        na_formula = '=' + '+'.join(na_parts) if na_parts else '=0'
        c = ws.cell(row=r, column=6, value=na_formula)
        c.font = VALUE_FONT
        c.border = THIN_BORDER

        # Total by level
        ws.cell(row=r, column=7,
                value=f'=SUM(B{r}:F{r})').font = VALUE_FONT
        ws.cell(row=r, column=7).border = THIN_BORDER

        r += 1

    return ws


# ── Helper to build a workbook with all common sheets ───────────────
def build_workbook(engagement_name, regulatory_basis, scope_desc,
                   assessment_covers, engagement_type_label,
                   domain_sheets_fn, scoping_note=None):
    """Build a complete workbook with methodology, scoping, planning,
    domain assessment sheets, Part D, reporting, and Part C."""
    wb = openpyxl.Workbook()
    create_methodology_sheet(wb, engagement_name, regulatory_basis,
                             scope_desc, assessment_covers, engagement_type_label)
    add_scoping_sheet(wb, engagement_name, scoping_note=scoping_note)
    add_planning_sheet(wb)
    domain_sheets_fn(wb)  # Add domain-specific assessment sheets
    add_part_d_sheet(wb)
    add_reporting_sheet(wb)
    add_part_c_sheet(wb)
    add_scoring_sheet(wb)
    return wb


# ═══════════════════════════════════════════════════════════════════
# CLOUD AWP — APPENDIX 10
# ═══════════════════════════════════════════════════════════════════

def add_cloud_sheets(wb):
    # ── Part A — Governance (7 areas) ──
    ws = wb.create_sheet(title='App 10 Part A - Governance')
    apply_col_widths(ws)
    r = 1
    write_header_row(ws, r)
    r += 1

    part_a = [
        ('Area 1 — Cloud Risk Management',
         'BNM expects board-level governance and a comprehensive cloud risk management framework integrated with the TRMF/CRF. The FI remains fully responsible for data protection regardless of cloud arrangements. Framework must articulate accountability and be periodically reviewed.',
         [
            ('DOMAIN', 'App10-A1', 'CLD-01', 'ORG', 'Cloud Risk Management',
             'Board governance and risk framework for cloud adoption'),
            ('SUB', 'App10-A1(a)', 'CLD-01.1', 'Board governance principles',
             '1. Obtain evidence of board promoting sound governance principles for cloud usage.\n2. Verify board/designated committee has approved cloud strategy and risk appetite.\n3. Check board receives periodic reporting on cloud risk posture.',
             'Board-approved cloud governance framework; Board/committee minutes; Cloud risk reports to board', 'Inspection'),
            ('SUB', 'App10-A1(b)', 'CLD-01.2', 'Cloud risk management framework integrated with TRMF/CRF',
             '1. Obtain cloud risk management framework.\n2. Verify integration with overall TRMF and CRF.\n3. Check framework covers risk identification, assessment, mitigation, monitoring for cloud.\n4. Verify alignment with enterprise risk management approach.',
             'Cloud risk management framework; TRMF/CRF integration evidence; Risk assessment methodology', 'Inspection'),
            ('SUB', 'App10-A1(c)', 'CLD-01.3', 'Framework tailored to service models with shared responsibility',
             '1. Verify framework addresses IaaS, PaaS, SaaS service models.\n2. Check shared responsibility model documented per service model.\n3. Confirm controls mapped to shared responsibility boundaries.\n4. Verify framework addresses deployment models (public, private, hybrid).',
             'Service model risk assessment; Shared responsibility matrix per CSP; Deployment model risk analysis', 'Inspection'),
            ('SUB', 'App10-A1(d)', 'CLD-01.4', 'FI responsible for data protection',
             '1. Verify FI acknowledges responsibility for customer data protection in cloud.\n2. Check data protection obligations documented regardless of cloud arrangement.\n3. Verify data classification controls applied to cloud-hosted data.\n4. Confirm data handling requirements flow down to CSP contracts.',
             'Data protection policy for cloud; Data classification in cloud; Contractual data protection provisions', 'Inspection'),
            ('SUB', 'App10-A1(e)', 'CLD-01.5', 'Onus on FI for customer information',
             '1. Verify FI retains full accountability for customer information security.\n2. Check customer notification/consent for cloud data hosting where required.\n3. Verify incident notification to customers covers cloud-related incidents.\n4. Confirm customer complaint handling covers cloud service issues.',
             'Customer information accountability policy; Customer notification records; Incident notification procedures', 'Inspection'),
            ('SUB', 'App10-A1(f)', 'CLD-01.6', 'Framework articulates board/senior management accountability',
             '1. Verify framework clearly assigns accountability to board and senior management.\n2. Check designated roles for cloud governance oversight.\n3. Verify escalation paths for cloud risk decisions.\n4. Confirm senior management performance includes cloud risk responsibilities.',
             'Accountability matrix; Cloud governance roles; Escalation procedures; Senior management KPIs', 'Inspection / Inquiry'),
            ('SUB', 'App10-A1(g)', 'CLD-01.7', 'Periodic review at least every 3 years',
             '1. Obtain review history of the cloud risk management framework.\n2. Verify reviews conducted at least every 3 years or upon material changes.\n3. Check review outcomes actioned and framework updated.\n4. Confirm review covers effectiveness of risk mitigation measures.',
             'Framework review history; Review reports; Action items and completion evidence', 'Inspection'),
         ]),
        ('Area 2 — Cloud Usage Policy',
         'BNM expects documented policies for cloud hosting criteria, a comprehensive asset inventory that includes cloud assets, and regular policy reviews. Without clear policies, cloud adoption decisions are inconsistent and uncontrolled.',
         [
            ('DOMAIN', 'App10-A2', 'CLD-02', 'ORG', 'Cloud Usage Policy',
             'Policies for cloud hosting criteria, asset inventory, and regular review'),
            ('SUB', 'App10-A2(a)', 'CLD-02.1', 'Policies for cloud hosting criteria',
             '1. Obtain cloud usage policy.\n2. Verify it defines criteria for what can/cannot be hosted in cloud.\n3. Check data classification requirements for cloud placement.\n4. Verify approval workflow for new cloud deployments.\n5. Confirm policy addresses service model selection criteria.',
             'Cloud usage policy; Data classification-to-cloud mapping; Cloud deployment approval workflow', 'Inspection'),
            ('SUB', 'App10-A2(b)', 'CLD-02.2', 'Asset inventory including cloud',
             '1. Obtain IT asset inventory.\n2. Verify cloud assets are included (accounts, subscriptions, services, data stores).\n3. Check inventory is accurate and up to date.\n4. Verify cloud asset discovery mechanisms in place.\n5. Confirm data classification of cloud-hosted assets recorded.',
             'IT asset inventory with cloud assets; Cloud account/subscription list; Asset discovery tool config', 'Inspection'),
            ('SUB', 'App10-A2(c)', 'CLD-02.3', 'Regular policy review every 3 years',
             '1. Obtain policy version history.\n2. Verify review cycle is at least every 3 years.\n3. Check policy updated to reflect changes in cloud landscape, regulation, or risk.\n4. Confirm policy review involves relevant stakeholders.',
             'Policy version history; Review schedule; Stakeholder sign-off records', 'Inspection'),
         ]),
        ('Area 3 — Due Diligence',
         'BNM requires risk-based due diligence before engaging a CSP. The FI must assess the CSP across multiple risk dimensions including financial stability, security, compliance, and operational resilience.',
         [
            ('DOMAIN', 'App10-A3', 'CLD-03', 'PLATFORM', 'Due Diligence',
             'Risk-based due diligence for CSP selection and ongoing assessment'),
            ('SUB', 'App10-A3', 'CLD-03.1', 'CSP due diligence assessment',
             '1. Obtain due diligence reports for each in-scope CSP.\n2. Verify coverage: financial viability, security capabilities, regulatory compliance, data residency, subcontracting, BCP.\n3. Check risk-based approach — higher scrutiny for critical workloads.\n4. Verify due diligence refreshed at least annually or upon material changes.\n5. Confirm due diligence findings inform CSP selection decision.',
             'Due diligence reports per CSP; Risk assessment methodology; Annual refresh evidence; CSP selection decision records', 'Inspection'),
         ]),
        ('Area 4 — CSP Certifications',
         'BNM expects the FI to seek assurance on CSP compliance through independent certifications and reports. This provides evidence that CSP controls have been independently validated.',
         [
            ('DOMAIN', 'App10-A4', 'CLD-04', 'PLATFORM', 'CSP Certifications',
             'Independent assurance on CSP compliance and control effectiveness'),
            ('SUB', 'App10-A4(a)', 'CLD-04.1', 'Assurance on compliance and action plans',
             '1. Verify FI seeks and obtains assurance on CSP compliance with relevant standards.\n2. Check CSP compliance certificates are current (ISO 27001, SOC 2, CSA STAR).\n3. Verify certificates cover the services and regions used by the FI.\n4. Check FI reviews CSP action plans for any identified gaps.',
             'CSP compliance certificates; Certificate scope verification; CSP action plans for gaps', 'Inspection'),
            ('SUB', 'App10-A4(b)', 'CLD-04.2', 'Independent external party reports',
             '1. Obtain current SOC 2 Type II reports per CSP.\n2. Verify report covers services used by the FI.\n3. Review exceptions or deficiencies noted in SOC reports.\n4. Check CUECs identified in SOC reports are implemented by the FI.\n5. Verify CSA STAR or equivalent additional reports obtained where available.',
             'SOC 2 Type II reports; CUEC implementation evidence; CSA STAR reports; FI review of report findings', 'Inspection'),
         ]),
        ('Area 5 — Contract Management',
         'BNM requires contracts to address cloud-specific and third-party risks including jurisdiction, regulatory compliance, and fourth-party service providers. Contractual provisions are the primary enforcement mechanism for cloud governance.',
         [
            ('DOMAIN', 'App10-A5', 'CLD-05', 'PLATFORM', 'Contract Management',
             'Cloud contract provisions for risk, jurisdiction, compliance, and fourth parties'),
            ('SUB', 'App10-A5(a)', 'CLD-05.1', 'Contracts address cloud/third party risks',
             '1. Review contracts with each in-scope CSP.\n2. Verify contracts address: right to audit, data ownership/portability, data residency, breach notification, termination/exit.\n3. Check contracts reviewed by legal and compliance.\n4. Verify BNM inspection access preserved.',
             'CSP contracts; Legal/compliance review records; Contract compliance checklist', 'Inspection'),
            ('SUB', 'App10-A5(b)(i)', 'CLD-05.2', 'Jurisdiction risks — data residency',
             '1. Identify data residency requirements per data classification.\n2. Verify contractual provisions specify permitted data locations.\n3. Check CSP configs prevent data replication to non-approved regions.\n4. Verify BNM approval obtained for data outside Malaysia if required.',
             'Data residency requirements; Contractual location provisions; CSP region config; BNM approval records', 'Inspection'),
            ('SUB', 'App10-A5(b)(ii)', 'CLD-05.3', 'Jurisdiction risks — legal and regulatory',
             '1. Assess legal/regulatory risks of hosting in foreign jurisdictions.\n2. Verify contract addresses governing law and dispute resolution.\n3. Check data protection laws of hosting jurisdiction assessed.\n4. Confirm cross-border data flow requirements met.',
             'Jurisdiction risk assessment; Legal opinion; Cross-border data flow assessment', 'Inspection'),
            ('SUB', 'App10-A5(b)(iii)', 'CLD-05.4', 'Jurisdiction risks — access by foreign authorities',
             '1. Assess risk of data access by foreign government/law enforcement.\n2. Verify notification provisions in contract for government access requests.\n3. Check encryption and key management mitigate unauthorised access risk.\n4. Verify FI awareness of extraterritorial data access laws.',
             'Foreign access risk assessment; Contract notification provisions; Encryption/key management for foreign access mitigation', 'Inspection'),
            ('SUB', 'App10-A5(c)', 'CLD-05.5', 'Potential impact and regulatory compliance',
             '1. Verify contract addresses potential impact of CSP failure or breach.\n2. Check financial penalties/SLA credits for service level failures.\n3. Confirm regulatory compliance obligations flow down to CSP.\n4. Verify FI can demonstrate compliance to BNM through CSP arrangements.',
             'Contract SLA and penalty provisions; Regulatory compliance flow-down clauses; Demonstrable compliance mechanism', 'Inspection'),
            ('SUB', 'App10-A5(d)(i)', 'CLD-05.6', 'Fourth party risks — sub-processor controls',
             '1. Identify CSP sub-processors/fourth parties.\n2. Verify contract requires CSP notification of sub-processor changes.\n3. Check FI approval rights for material sub-processor changes.\n4. Verify sub-processor security requirements equivalent to CSP obligations.',
             'Sub-processor list; Contract sub-processor provisions; Notification/approval records', 'Inspection'),
            ('SUB', 'App10-A5(d)(ii)', 'CLD-05.7', 'Fourth party risks — oversight and concentration',
             '1. Assess concentration risk from key fourth parties.\n2. Verify CSP oversight extends to critical fourth parties.\n3. Check FI can obtain assurance on fourth party controls.\n4. Confirm fourth party failure scenarios addressed in BCP.',
             'Fourth party concentration analysis; CSP fourth party oversight evidence; BCP fourth party scenarios', 'Inspection'),
         ]),
        ('Area 6 — Oversight over CSPs',
         'BNM requires continuous monitoring mechanisms, defined key responsibilities, periodic assessments, and prompt risk reassessment upon material changes. Ongoing oversight ensures cloud risk remains within tolerance.',
         [
            ('DOMAIN', 'App10-A6', 'CLD-06', 'ORG', 'Oversight over CSPs',
             'Continuous monitoring, periodic assessments, and prompt reassessment on changes'),
            ('SUB', 'App10-A6(a)', 'CLD-06.1', 'Continuous monitoring mechanism',
             '1. Verify continuous monitoring of CSP performance and security posture.\n2. Check monitoring covers SLA compliance, incident notifications, security advisories.\n3. Verify automated monitoring tools deployed (CSPM, SLA dashboards).\n4. Confirm monitoring results reported to management.',
             'CSP monitoring framework; Monitoring tool config; SLA dashboards; Management reports', 'Inspection'),
            ('SUB', 'App10-A6(b)', 'CLD-06.2', 'Key responsibilities for monitoring',
             '1. Verify defined roles and responsibilities for CSP oversight.\n2. Check monitoring responsibilities assigned to named individuals/teams.\n3. Verify escalation procedures for CSP issues.\n4. Confirm oversight activities are documented.',
             'CSP oversight RACI; Named role assignments; Escalation procedures; Oversight activity records', 'Inspection / Inquiry'),
            ('SUB', 'App10-A6(c)', 'CLD-06.3', 'Periodic assessments',
             '1. Verify periodic CSP assessments conducted (at least annual).\n2. Check assessments cover security, performance, compliance, and contract adherence.\n3. Review most recent CSP assessment report.\n4. Confirm findings tracked and remediated.',
             'CSP assessment schedule; Assessment reports; Finding tracker; Remediation evidence', 'Inspection'),
            ('SUB', 'App10-A6(d)', 'CLD-06.4', 'Prompt review upon material changes',
             '1. Verify process for reassessing CSP risk upon material changes.\n2. Check triggers defined (CSP M&A, major incidents, regulation changes, service changes).\n3. Review any recent trigger events and confirm reassessment was performed.\n4. Verify reassessment outcomes documented.',
             'Material change trigger process; Trigger event log; Reassessment records', 'Inspection'),
         ]),
        ('Area 7 — Skilled Personnel',
         'BNM requires internal resources with cloud expertise, cross-functional involvement (not purely IT), equipped internal audit, adequate training, and consequence management. Without skilled personnel, the FI cannot independently govern cloud risk.',
         [
            ('DOMAIN', 'App10-A7', 'CLD-07', 'ORG', 'Skilled Personnel',
             'Internal cloud expertise, training, audit capability, and consequence management'),
            ('SUB', 'App10-A7(a)(i)', 'CLD-07.1', 'Internal resources — cloud expertise',
             '1. Review cloud team structure and headcount.\n2. Verify cloud certifications held by key personnel (CCSP, CCSK, AWS/Azure/GCP certs).\n3. Check skills gap analysis conducted.\n4. Verify FI can independently assess CSP configurations.',
             'Cloud team org chart; Certification records; Skills gap analysis', 'Inspection / Inquiry'),
            ('SUB', 'App10-A7(a)(ii)', 'CLD-07.2', 'Internal resources — succession planning',
             '1. Verify succession planning for critical cloud roles.\n2. Check cross-training in place.\n3. Confirm knowledge documentation for key processes.\n4. Verify contingency arrangements for key personnel departure.',
             'Succession plans; Cross-training records; Knowledge base; Contingency arrangements', 'Inspection'),
            ('SUB', 'App10-A7(b)', 'CLD-07.3', 'Not purely IT responsibility',
             '1. Verify cloud governance involves risk management, compliance, legal, and business functions.\n2. Check cloud governance committee/forum includes non-IT representatives.\n3. Verify business stakeholders participate in cloud risk decisions.\n4. Confirm risk and compliance functions have cloud oversight roles.',
             'Governance committee membership; Meeting minutes showing cross-functional participation; RACI showing non-IT roles', 'Inspection'),
            ('SUB', 'App10-A7(c)', 'CLD-07.4', 'Equip internal audit',
             '1. Verify internal audit has cloud audit capability.\n2. Check audit plan includes cloud-specific audits.\n3. Verify internal auditors have relevant cloud training/certifications.\n4. Confirm audit methodology covers cloud-specific risks.',
             'Internal audit plan; Cloud audit reports; Auditor cloud certifications; Cloud audit methodology', 'Inspection'),
            ('SUB', 'App10-A7(d)', 'CLD-07.5', 'Adequate training',
             '1. Obtain training plan for cloud-related staff.\n2. Verify training covers security, operations, and compliance.\n3. Check training records for completeness.\n4. Verify training currency (refreshed at least annually).',
             'Training plan; Training records; Training content; Training completion rates', 'Inspection'),
            ('SUB', 'App10-A7(e)', 'CLD-07.6', 'Consequence management',
             '1. Verify consequence management framework covers cloud-related non-compliance.\n2. Check accountability for cloud security breaches defined.\n3. Verify enforcement actions documented.\n4. Confirm awareness of consequences communicated to staff.',
             'Consequence management framework; Accountability provisions; Enforcement records; Awareness materials', 'Inspection'),
         ]),
    ]

    r = write_assessment_rows(ws, r, part_a)

    # ── Part B — Design & Controls (14 areas) ──
    ws = wb.create_sheet(title='App 10 Part B - Controls')
    apply_col_widths(ws)
    r = 1
    write_header_row(ws, r)
    r += 1

    part_b = [
        ('B1 — Cloud Architecture',
         'Cloud architecture must support resilience, security, and scalability. Expect multi-AZ design, proper segmentation, well-architected review. Poor architecture creates systemic risk.',
         [
            ('DOMAIN', 'App10-B1', 'CLD-08', 'PLATFORM', 'Cloud Architecture',
             'Architecture design, multi-AZ, segmentation, and landing zone'),
            ('SUB', 'App10-B1(a)', 'CLD-08.1', 'Architecture design principles',
             '1. Obtain cloud architecture design documents.\n2. Verify adherence to well-architected framework principles.\n3. Check architecture reviewed by qualified cloud architects.',
             'Cloud architecture docs; Well-architected review reports; Architect review records', 'Inspection'),
            ('SUB', 'App10-B1(b)', 'CLD-08.2', 'Multi-AZ/multi-region for critical workloads',
             '1. Verify multi-AZ or multi-region deployment for critical workloads.\n2. Check failover mechanisms tested.\n3. Confirm RTO/RPO met by architecture design.',
             'Multi-AZ/region deployment evidence; Failover test results; RTO/RPO validation', 'Inspection'),
            ('SUB', 'App10-B1(c)', 'CLD-08.3', 'Network architecture and segmentation',
             '1. Review VPC/VNET design and subnet segmentation.\n2. Verify security groups and NACLs follow least-privilege.\n3. Check isolation between workloads and environments.',
             'VPC/VNET design; Security group configs; NACL rules; Environment isolation evidence', 'Inspection'),
            ('SUB', 'App10-B1(d)', 'CLD-08.4', 'Landing zone and account structure',
             '1. Verify landing zone / account structure design.\n2. Check separation of concerns (production, staging, development, sandbox).\n3. Verify guardrails and SCPs applied.\n4. Confirm account provisioning follows approved process.',
             'Landing zone documentation; Account structure; SCP/guardrail configs; Account provisioning process', 'Inspection'),
            ('SUB', 'App10-B1(e)', 'CLD-08.5', 'API and service endpoint security',
             '1. Review API gateway and service endpoint configurations.\n2. Verify authentication and authorisation for all endpoints.\n3. Check rate limiting and throttling.\n4. Confirm private endpoints used where appropriate.',
             'API gateway config; Endpoint auth config; Rate limiting; Private endpoint deployment', 'Inspection'),
            ('SUB', 'App10-B1(f)', 'CLD-08.6', 'Architecture documentation and currency',
             '1. Verify architecture documentation is current and version-controlled.\n2. Check diagrams reflect actual deployment.\n3. Confirm change management for architecture changes.\n4. Verify architecture reviewed at least annually.',
             'Architecture version history; Diagram accuracy verification; Change records; Annual review evidence', 'Inspection'),
         ]),
        ('B2 — Cloud Application Delivery',
         'Application delivery must be secure and governed. Expect secure CI/CD pipelines, security gates, controlled deployments. Uncontrolled application delivery introduces vulnerabilities.',
         [
            ('DOMAIN', 'App10-B2', 'CLD-09', 'PLATFORM', 'Cloud Application Delivery',
             'Secure CI/CD, security gates, and deployment governance'),
            ('SUB', 'App10-B2(a)', 'CLD-09.1', 'CI/CD pipeline security',
             '1. Review CI/CD pipeline architecture.\n2. Verify pipeline configs version-controlled and access-restricted.\n3. Check secrets not hardcoded in pipeline configurations.\n4. Confirm pipeline runs in secure, isolated environments.',
             'CI/CD architecture; Pipeline access controls; Secrets management; Pipeline isolation', 'Inspection'),
            ('SUB', 'App10-B2(b)(i)', 'CLD-09.2', 'Security gates — static and dynamic analysis',
             '1. Verify SAST integrated into CI/CD pipeline.\n2. Check DAST performed before production deployment.\n3. Confirm security gate failures block deployment.\n4. Verify SCA for third-party dependencies.',
             'SAST integration config; DAST reports; Gate failure records; SCA scan results', 'Inspection'),
            ('SUB', 'App10-B2(b)(ii)', 'CLD-09.3', 'Security gates — container and IaC scanning',
             '1. Verify container image scanning before deployment.\n2. Check IaC templates scanned for misconfigurations.\n3. Confirm base images from approved registries only.\n4. Verify vulnerability thresholds enforced.',
             'Container scanning config; IaC scanning tool; Approved registry list; Vulnerability threshold policy', 'Inspection'),
            ('SUB', 'App10-B2(c)(i)', 'CLD-09.4', 'Deployment governance — approval workflows',
             '1. Verify production deployment requires approval.\n2. Check approval workflow documented and enforced.\n3. Confirm emergency deployment procedures exist.\n4. Verify deployment rollback capability.',
             'Deployment approval workflow; Approval records; Emergency procedures; Rollback evidence', 'Inspection'),
            ('SUB', 'App10-B2(c)(ii)', 'CLD-09.5', 'Deployment governance — environment separation',
             '1. Verify environment separation (dev, staging, production).\n2. Check production data not used in non-production without masking.\n3. Confirm separate credentials per environment.\n4. Verify deployment promotion process controls.',
             'Environment architecture; Data masking evidence; Per-environment credentials; Promotion controls', 'Inspection'),
            ('SUB', 'App10-B2(c)(iii)', 'CLD-09.6', 'Deployment governance — release management',
             '1. Verify release management process.\n2. Check release notes and change documentation.\n3. Confirm post-deployment validation.\n4. Verify rollback executed successfully when needed.',
             'Release management process; Release notes; Post-deployment validation; Rollback records', 'Inspection'),
         ]),
        ('B3 — Virtualisation/Containerisation',
         'Virtualisation and containerisation introduce shared-resource risks. Expect hypervisor hardening, container runtime security, image governance, and workload isolation. Escape vulnerabilities can cross tenant boundaries.',
         [
            ('DOMAIN', 'App10-B3', 'CLD-10', 'PLATFORM', 'Virtualisation/Containerisation',
             'Hypervisor security, container governance, runtime controls, and workload isolation'),
            ('SUB', 'App10-B3(a)', 'CLD-10.1', 'Hypervisor and host security',
             '1. Verify hypervisor hardening per vendor/CIS benchmarks.\n2. Check host OS patching and vulnerability management.\n3. Confirm management plane access restricted and monitored.',
             'Hypervisor hardening evidence; CIS benchmark compliance; Patch status; Management access controls', 'Inspection'),
            ('SUB', 'App10-B3(b)(i)', 'CLD-10.2', 'Container image governance',
             '1. Verify container images sourced from approved registries.\n2. Check image scanning for vulnerabilities before deployment.\n3. Confirm base images maintained and updated.\n4. Verify image signing and integrity verification.',
             'Approved registry list; Image scanning reports; Base image maintenance; Image signing config', 'Inspection'),
            ('SUB', 'App10-B3(b)(ii)', 'CLD-10.3', 'Container runtime security',
             '1. Verify containers run as non-root.\n2. Check read-only file systems enforced where possible.\n3. Confirm resource limits set (CPU, memory).\n4. Verify runtime security monitoring deployed.',
             'Container security policy; Runtime config; Resource limits; Runtime monitoring', 'Inspection'),
            ('SUB', 'App10-B3(b)(iii)', 'CLD-10.4', 'Container network policies',
             '1. Verify pod/container network policies enforce segmentation.\n2. Check default deny network policy.\n3. Confirm inter-service communication authenticated.\n4. Verify service mesh security where deployed.',
             'Network policies; Default deny evidence; Service authentication; Service mesh config', 'Inspection'),
            ('SUB', 'App10-B3(b)(iv)', 'CLD-10.5', 'Orchestrator hardening',
             '1. Verify orchestrator (Kubernetes, ECS, etc.) hardened per CIS benchmarks.\n2. Check RBAC configured with least privilege.\n3. Confirm API server access restricted.\n4. Verify secrets management for orchestrator.',
             'CIS benchmark compliance; RBAC config; API server access controls; Secrets management', 'Inspection'),
            ('SUB', 'App10-B3(b)(v)', 'CLD-10.6', 'Container logging and monitoring',
             '1. Verify container logs captured and centralised.\n2. Check security events from containers forwarded to SIEM.\n3. Confirm container anomaly detection.\n4. Verify log retention for containers.',
             'Container logging config; SIEM integration; Anomaly detection; Log retention', 'Inspection'),
            ('SUB', 'App10-B3(b)(vi)', 'CLD-10.7', 'Workload isolation',
             '1. Verify workload isolation between tenants/environments.\n2. Check namespace isolation in orchestrator.\n3. Confirm resource quotas prevent noisy neighbour.\n4. Verify sensitive workloads on dedicated infrastructure where required.',
             'Isolation architecture; Namespace config; Resource quotas; Dedicated infrastructure evidence', 'Inspection'),
         ]),
        ('B4 — Change Management',
         'Cloud changes must be controlled to prevent misconfigurations. Expect change approval, risk assessment, IaC governance, and configuration drift detection.',
         [
            ('DOMAIN', 'App10-B4', 'CLD-11', 'PLATFORM', 'Change Management',
             'Cloud change controls, IaC governance, and drift detection'),
            ('SUB', 'App10-B4(a)', 'CLD-11.1', 'Cloud change management procedure',
             '1. Review cloud change management procedure.\n2. Sample 10 recent cloud changes.\n3. Verify each has: request, risk assessment, approval, implementation, post-implementation review.\n4. Check security-critical changes (IAM, network, encryption) have enhanced approval.',
             'Change management procedure; 10 sampled change records; Enhanced approval evidence', 'Inspection'),
            ('SUB', 'App10-B4(b)', 'CLD-11.2', 'Infrastructure as Code governance',
             '1. Review IaC governance (code review, version control, approval).\n2. Check IaC templates scanned for misconfigurations.\n3. Verify IaC state files stored securely.\n4. Confirm IaC changes follow standard change process.',
             'IaC repository; IaC scanning tool; State file storage; Change process for IaC', 'Inspection'),
            ('SUB', 'App10-B4(c)', 'CLD-11.3', 'Configuration drift detection',
             '1. Verify configuration drift detection deployed.\n2. Check drift alerts trigger remediation.\n3. Confirm baseline configurations defined per resource type.\n4. Verify emergency and rollback change procedures.',
             'Drift detection tool; Drift alert records; Baseline configs; Emergency change procedures', 'Inspection'),
         ]),
        ('B5 — Backup and Recovery',
         'Cloud backup and recovery must be tested, not assumed. Expect defined backup policies, cross-region replication, tested restores, and defined RTO/RPO per workload. Cloud availability does not substitute for tested DR.',
         [
            ('DOMAIN', 'App10-B5', 'CLD-12', 'WORKLOAD', 'Backup and Recovery',
             'Backup strategy, cross-region replication, restore testing, and RTO/RPO validation'),
            ('SUB', 'App10-B5(a)(i)', 'CLD-12.1', 'Backup policy and schedule',
             '1. Obtain cloud backup policy.\n2. Verify backup schedule covers all critical data stores.\n3. Check backup frequency aligned with RPO requirements.\n4. Confirm backup policy covers all cloud services (databases, object storage, file systems).',
             'Cloud backup policy; Backup schedule; RPO alignment evidence; Service coverage', 'Inspection'),
            ('SUB', 'App10-B5(a)(ii)', 'CLD-12.2', 'Backup encryption and access controls',
             '1. Verify backups encrypted at rest.\n2. Check backup access restricted to authorised personnel.\n3. Confirm backup key management follows organisational standards.\n4. Verify backup immutability (write-once) for critical data.',
             'Backup encryption config; Access controls; Key management; Immutability config', 'Inspection'),
            ('SUB', 'App10-B5(a)(iii)', 'CLD-12.3', 'Cross-region backup replication',
             '1. Verify cross-region replication for critical backups.\n2. Check replication targets comply with data residency requirements.\n3. Confirm replication lag monitored.\n4. Verify data integrity of replicated backups.',
             'Replication config; Data residency compliance; Replication monitoring; Integrity checks', 'Inspection'),
            ('SUB', 'App10-B5(b)', 'CLD-12.4', 'Backup monitoring',
             '1. Verify backup job monitoring in place.\n2. Check failed backups trigger alerts.\n3. Confirm backup success rates tracked.\n4. Verify backup capacity planning.',
             'Backup monitoring config; Alert records; Success rate reports; Capacity planning', 'Inspection'),
            ('SUB', 'App10-B5(c)(i)', 'CLD-12.5', 'Restore testing',
             '1. Verify restore testing conducted at least annually.\n2. Check restore tests cover representative workloads.\n3. Confirm restore completed within RTO.\n4. Verify data integrity validated after restore.',
             'Restore test plan; Test results; RTO achievement; Data integrity validation', 'Inspection'),
            ('SUB', 'App10-B5(c)(ii)', 'CLD-12.6', 'DR failover testing',
             '1. Verify cloud DR/failover tested (cross-region or cross-AZ).\n2. Check failover tests simulate realistic failure scenarios.\n3. Confirm failover completed within defined timeframes.\n4. Verify test findings remediated.',
             'DR failover test plan; Test results; Failover timelines; Remediation records', 'Inspection'),
            ('SUB', 'App10-B5(d)', 'CLD-12.7', 'RTO/RPO definitions and validation',
             '1. Verify RTO/RPO defined per critical cloud workload.\n2. Check RTO/RPO approved by business owners.\n3. Confirm RTO/RPO validated through testing.\n4. Verify CSP capabilities support defined RTO/RPO.',
             'RTO/RPO definitions; Business owner approval; Validation test results; CSP capability assessment', 'Inspection'),
            ('SUB', '', 'CLD-12.8', 'CSP outage response procedures',
             '1. Verify CSP outage response procedures documented.\n2. Check procedures cover communication, escalation, and workaround.\n3. Confirm lessons learned from past CSP outages reviewed.\n4. Verify integration with overall incident management.',
             'CSP outage procedures; Communication plan; Lessons learned; Incident management integration', 'Inspection'),
         ]),
        ('B6 — Interoperability/Portability',
         'Data and workload portability reduce vendor lock-in risk. Expect data export capabilities, standard formats, API compatibility assessment. Without portability, the FI is captive to the CSP.',
         [
            ('DOMAIN', 'App10-B6', 'CLD-13', 'PLATFORM', 'Interoperability/Portability',
             'Data portability, standard formats, and vendor lock-in mitigation'),
            ('SUB', 'App10-B6(a)', 'CLD-13.1', 'Data portability assessment',
             '1. Verify data portability assessed for each critical CSP service.\n2. Check data can be exported in standard, open formats.\n3. Confirm export mechanisms tested.\n4. Verify data completeness and integrity after export.',
             'Data portability assessment; Export format documentation; Export test results; Integrity verification', 'Inspection'),
            ('SUB', 'App10-B6(b)', 'CLD-13.2', 'Application portability',
             '1. Assess application dependencies on CSP-specific services.\n2. Verify abstraction layers used where feasible.\n3. Check containerisation and standardised APIs reduce lock-in.\n4. Confirm portability considerations in architecture decisions.',
             'CSP dependency analysis; Abstraction layer documentation; Architecture decision records', 'Inspection'),
            ('SUB', 'App10-B6(c)', 'CLD-13.3', 'Multi-cloud or hybrid capability',
             '1. Verify multi-cloud or hybrid deployment capability where appropriate.\n2. Check interoperability between cloud environments.\n3. Confirm management tooling supports multi-environment.\n4. Verify network connectivity between environments.',
             'Multi-cloud architecture; Interoperability evidence; Management tooling; Network connectivity', 'Inspection'),
            ('SUB', 'App10-B6(d)', 'CLD-13.4', 'API compatibility and standardisation',
             '1. Verify APIs follow industry standards (REST, OpenAPI).\n2. Check API versioning strategy.\n3. Confirm API documentation maintained.\n4. Verify API-level portability assessed.',
             'API standards; Versioning strategy; API documentation; Portability assessment', 'Inspection'),
            ('SUB', 'App10-B6(e)', 'CLD-13.5', 'Vendor lock-in risk assessment',
             '1. Verify vendor lock-in risk formally assessed.\n2. Check lock-in mitigation strategies documented.\n3. Confirm risk accepted at appropriate level.\n4. Verify periodic reassessment.',
             'Lock-in risk assessment; Mitigation strategies; Risk acceptance records; Reassessment schedule', 'Inspection'),
         ]),
        ('B7 — Exit Strategy',
         'BNM requires a documented and tested exit strategy. Expect trigger events, data retrieval procedures, migration approach, timeline, and cost estimates. Without exit planning, the FI cannot leave a CSP.',
         [
            ('DOMAIN', 'App10-B7', 'CLD-14', 'PLATFORM', 'Exit Strategy',
             'Exit planning, trigger events, data retrieval, migration approach, and cost estimates'),
            ('SUB', 'App10-B7(a)(i)', 'CLD-14.1', 'Exit strategy — trigger events',
             '1. Obtain exit strategy for each critical CSP engagement.\n2. Verify trigger events defined (CSP failure, contract expiry, regulatory direction, strategic change).\n3. Check triggers actively monitored.',
             'Exit strategy document; Trigger event definitions; Monitoring evidence', 'Inspection'),
            ('SUB', 'App10-B7(a)(ii)', 'CLD-14.2', 'Exit strategy — data retrieval',
             '1. Verify data retrieval procedures documented.\n2. Check data export formats are standard and tested.\n3. Confirm data completeness verification procedure.\n4. Verify data deletion from CSP post-exit.',
             'Data retrieval procedures; Export format docs; Completeness verification; Deletion procedures', 'Inspection'),
            ('SUB', 'App10-B7(a)(iii)', 'CLD-14.3', 'Exit strategy — migration approach',
             '1. Verify migration approach documented (lift-and-shift, re-platform, re-architect).\n2. Check resource and skill requirements identified.\n3. Confirm timeline is realistic.\n4. Verify dependencies identified.',
             'Migration approach document; Resource plan; Timeline; Dependency analysis', 'Inspection'),
            ('SUB', 'App10-B7(a)(iv)', 'CLD-14.4', 'Exit strategy — communication plan',
             '1. Verify communication plan for exit scenario.\n2. Check stakeholder notification procedures.\n3. Confirm regulatory notification requirements addressed.\n4. Verify customer impact assessment.',
             'Communication plan; Stakeholder notification; Regulatory notification; Customer impact assessment', 'Inspection'),
            ('SUB', 'App10-B7(b)(i)', 'CLD-14.5', 'Contractual exit provisions — transition period',
             '1. Verify contract provides for adequate transition period.\n2. Check CSP cooperation obligations during exit.\n3. Confirm service continuity during transition.\n4. Verify transition support fees understood.',
             'Contract exit clauses; Transition period provisions; Service continuity terms; Fee schedule', 'Inspection'),
            ('SUB', 'App10-B7(b)(ii)', 'CLD-14.6', 'Contractual exit provisions — data return',
             '1. Verify contractual right to data return in usable format.\n2. Check data destruction obligations on CSP post-return.\n3. Confirm destruction certification from CSP.\n4. Verify timeline for data return.',
             'Data return contract terms; Destruction obligations; Destruction certification template; Return timeline', 'Inspection'),
            ('SUB', 'App10-B7(b)(iii)', 'CLD-14.7', 'Exit strategy review and testing',
             '1. Verify exit strategy reviewed within 12 months.\n2. Check exit feasibility assessed.\n3. Confirm migration feasibility tested or simulated.\n4. Review cost estimates for exit scenarios.',
             'Exit strategy review records; Feasibility assessment; Migration test/simulation; Cost estimates', 'Inspection'),
            ('SUB', 'App10-B7(b)(iv)', 'CLD-14.8', 'Alternative provider identification',
             '1. Verify alternative CSPs identified for critical services.\n2. Check compatibility assessment with alternative providers.\n3. Confirm alternative providers can meet regulatory requirements.\n4. Verify transition effort estimated.',
             'Alternative provider analysis; Compatibility assessment; Regulatory compliance check; Transition estimates', 'Inspection'),
         ]),
        ('B8 — Cryptographic Key Management',
         'Key management determines who can access encrypted data. Expect key ownership strategy, rotation, HSM usage, and key lifecycle management. Weak key management renders encryption ineffective.',
         [
            ('DOMAIN', 'App10-B8', 'CLD-15', 'PLATFORM', 'Cryptographic Key Management',
             'Key ownership, rotation, HSM, and lifecycle management'),
            ('SUB', 'App10-B8(a)', 'CLD-15.1', 'Key ownership model',
             '1. Identify key ownership model per service (CSP-managed, CMK, BYOK, HYOK).\n2. Verify key ownership decision based on data classification.\n3. Check FI maintains control over keys for sensitive data.\n4. Confirm key ownership documented.',
             'Key ownership matrix; Data classification-to-key mapping; Key control documentation', 'Inspection'),
            ('SUB', 'App10-B8(b)', 'CLD-15.2', 'Key rotation and lifecycle',
             '1. Verify key rotation schedule (at least annual for symmetric keys).\n2. Check key lifecycle management (generation, distribution, storage, rotation, revocation, destruction).\n3. Confirm key version management.\n4. Verify automated rotation where possible.',
             'Key rotation schedule; Lifecycle management procedures; Key version tracking; Automation config', 'Inspection'),
            ('SUB', 'App10-B8(c)', 'CLD-15.3', 'HSM and secure key storage',
             '1. Verify HSM usage for key storage where appropriate.\n2. Check HSM compliance certifications (FIPS 140-2 Level 3+).\n3. Confirm key access controls restrict to authorised services/personnel.\n4. Verify key backup and recovery procedures.',
             'HSM config; FIPS certification; Key access controls; Key backup procedures', 'Inspection'),
            ('SUB', 'App10-B8(d)', 'CLD-15.4', 'Encryption standards and algorithms',
             '1. Verify encryption algorithms meet minimum standards (AES-256, RSA-2048+).\n2. Check no deprecated algorithms in use.\n3. Confirm encryption applied at rest and in transit for all sensitive data.\n4. Verify crypto agility — ability to update algorithms when needed.',
             'Encryption standards policy; Algorithm inventory; At-rest/in-transit encryption audit; Crypto agility assessment', 'Inspection'),
         ]),
        ('B9 — Access Controls',
         'Cloud IAM is the primary control plane. Expect MFA, least privilege, service account management, privileged access controls. Weak IAM is the most common cloud compromise vector.',
         [
            ('DOMAIN', 'App10-B9', 'CLD-16', 'PLATFORM', 'Access Controls',
             'IAM, MFA, least privilege, service accounts, and privileged access management'),
            ('SUB', 'App10-B9(a)(i)', 'CLD-16.1', 'IAM architecture and federation',
             '1. Review cloud IAM architecture.\n2. Verify identity federation with enterprise IdP.\n3. Check SSO integration.\n4. Confirm centralised identity management.',
             'IAM architecture; Federation config; SSO integration; Centralised management evidence', 'Inspection'),
            ('SUB', 'App10-B9(a)(ii)', 'CLD-16.2', 'MFA enforcement',
             '1. Verify MFA enforced for all cloud console and API access.\n2. Check MFA method strength (hardware tokens, FIDO2 preferred).\n3. Confirm MFA not bypassable.\n4. Verify MFA for root/break-glass accounts.',
             'MFA enforcement config; MFA method assessment; Bypass prevention; Root account MFA', 'Inspection'),
            ('SUB', 'App10-B9(a)(iii)', 'CLD-16.3', 'Least privilege enforcement',
             '1. Review IAM policies for overly permissive roles.\n2. Check for wildcard permissions.\n3. Verify permission boundaries applied.\n4. Confirm unused permissions removed (access advisor).',
             'IAM policy review; Wildcard permission audit; Permission boundaries; Access advisor reports', 'Inspection'),
            ('SUB', 'App10-B9(a)(iv)', 'CLD-16.4', 'Service account and API key management',
             '1. Obtain service account inventory.\n2. Verify service accounts follow least privilege.\n3. Check API key rotation enforced.\n4. Confirm unused service accounts and keys disabled.\n5. Verify secrets not embedded in code.',
             'Service account inventory; Permission review; Key rotation config; Unused account audit; Code scanning for secrets', 'Inspection'),
            ('SUB', 'App10-B9(a)(v)', 'CLD-16.5', 'Access reviews',
             '1. Verify periodic access reviews for cloud IAM (at least quarterly for privileged).\n2. Check reviews result in access modifications.\n3. Confirm stale/orphan accounts identified and removed.\n4. Verify review evidence documented.',
             'Access review schedule; Review records; Access modification evidence; Stale account remediation', 'Inspection'),
            ('SUB', 'App10-B9(b)', 'CLD-16.6', 'Privileged access management',
             '1. Verify privileged/admin access limited and separately managed.\n2. Check PAM solution deployed for cloud admin access.\n3. Confirm session recording for privileged sessions.\n4. Verify just-in-time access for administrative tasks.',
             'Privileged access inventory; PAM config; Session recording; JIT access config', 'Inspection'),
            ('SUB', 'App10-B9(c)', 'CLD-16.7', 'Break-glass / emergency access',
             '1. Verify break-glass procedures documented.\n2. Check break-glass accounts secured (hardware MFA, vault-managed).\n3. Confirm break-glass usage triggers alerts and review.\n4. Verify break-glass tested at least annually.',
             'Break-glass procedures; Account security config; Alert config; Test records', 'Inspection'),
            ('SUB', 'App10-B9(d)(i)', 'CLD-16.8', 'Root account security',
             '1. Verify root/owner accounts not used for daily operations.\n2. Check root access secured with hardware MFA.\n3. Confirm root usage monitored and alerted.\n4. Verify root credentials stored securely.',
             'Root account usage policy; Hardware MFA config; Root usage alerts; Credential storage', 'Inspection'),
            ('SUB', 'App10-B9(d)(ii)', 'CLD-16.9', 'Cross-account access controls',
             '1. Verify cross-account access governed by policy.\n2. Check trust relationships reviewed periodically.\n3. Confirm cross-account roles follow least privilege.\n4. Verify logging of cross-account access.',
             'Cross-account access policy; Trust relationship review; Role permissions; Cross-account logging', 'Inspection'),
         ]),
        ('B10 — Cybersecurity Operations',
         'Cloud security operations require adapted monitoring, detection, and response capabilities. Expect cloud-native logging, SIEM integration, CSPM, and cloud-specific SOC coverage.',
         [
            ('DOMAIN', 'App10-B10', 'CLD-17', 'PLATFORM', 'Cybersecurity Operations',
             'Cloud security monitoring, CSPM, SIEM integration, and SOC coverage'),
            ('SUB', 'App10-B10(a)', 'CLD-17.1', 'Cloud audit logging',
             '1. Verify cloud-native audit logging enabled (CloudTrail, Activity Log, etc.).\n2. Confirm logs cover all regions and accounts.\n3. Check log integrity protection (tamper-proof).\n4. Verify log retention aligned with regulatory requirements.',
             'Cloud audit log config; Region/account coverage; Log integrity protection; Retention config', 'Inspection'),
            ('SUB', 'App10-B10(b)(i)', 'CLD-17.2', 'SIEM integration and alerting',
             '1. Confirm centralised log aggregation with SIEM.\n2. Check alert rules for security-relevant cloud events.\n3. Verify SOC coverage extends to cloud events.\n4. Confirm cloud-specific detection rules (IAM changes, data exposure, config changes).',
             'SIEM integration; Alert rules; SOC cloud coverage; Cloud detection rules', 'Inspection'),
            ('SUB', 'App10-B10(b)(ii)', 'CLD-17.3', 'CSPM deployment and findings',
             '1. Verify CSPM deployed to detect misconfigurations.\n2. Review CSPM findings and remediation status.\n3. Check critical findings remediated within defined SLA.\n4. Confirm CSPM covers all cloud accounts and services.',
             'CSPM config; Findings report; Remediation tracker; Coverage assessment', 'Inspection'),
            ('SUB', 'App10-B10(c)(i)', 'CLD-17.4', 'Cloud incident response',
             '1. Review IR plan for cloud-specific scenarios (account compromise, data breach, CSP outage).\n2. Verify procedures account for shared responsibility model.\n3. Check CSP escalation contacts documented.\n4. Verify IR exercises include cloud scenarios.',
             'Cloud IR procedures; CSP escalation contacts; IR exercise records', 'Inspection / Inquiry'),
            ('SUB', 'App10-B10(c)(ii)', 'CLD-17.5', 'Cloud forensic capability',
             '1. Verify forensic capability in cloud (snapshot, log preservation, evidence collection).\n2. Check forensic procedures documented.\n3. Confirm forensic tools available.\n4. Verify chain of custody procedures for cloud evidence.',
             'Forensic procedures; Tool inventory; Evidence collection capability; Chain of custody', 'Inspection'),
         ]),
        ('B11 — DDoS',
         'DDoS protection for cloud-hosted services is essential. Expect L3/L4 and L7 protection, auto-scaling, and incident response procedures.',
         [
            ('DOMAIN', 'App10-B11', 'CLD-18', 'PLATFORM', 'DDoS Protection',
             'Layered DDoS protection, auto-scaling, and response procedures'),
            ('SUB', 'App10-B11(a)', 'CLD-18.1', 'DDoS protection architecture',
             '1. Verify DDoS protection enabled (L3/L4 and L7).\n2. Check CSP-native DDoS protection activated.\n3. Confirm WAF deployed for application-layer protection.\n4. Verify auto-scaling configured to absorb traffic spikes.',
             'DDoS protection config; CSP DDoS service; WAF config; Auto-scaling config', 'Inspection'),
            ('SUB', 'App10-B11(b)', 'CLD-18.2', 'DDoS response and testing',
             '1. Verify DDoS response procedures documented.\n2. Check DDoS testing/simulation conducted.\n3. Confirm CSP DDoS support engagement process understood.\n4. Verify monitoring detects DDoS attacks promptly.',
             'DDoS response procedures; Test/simulation results; CSP support process; DDoS monitoring', 'Inspection'),
         ]),
        ('B12 — DLP',
         'Data Loss Prevention controls prevent exfiltration of sensitive data from cloud environments. Expect DLP policies, monitoring, and response procedures.',
         [
            ('DOMAIN', 'App10-B12', 'CLD-19', 'PLATFORM', 'Data Loss Prevention',
             'DLP policies, monitoring, and exfiltration controls'),
            ('SUB', 'App10-B12(a)', 'CLD-19.1', 'DLP policy and controls',
             '1. Verify DLP policies cover cloud storage, email, endpoints.\n2. Check sensitive data patterns defined and monitored.\n3. Confirm public access to cloud storage blocked.\n4. Verify data sharing controls enforced.',
             'DLP policy config; Sensitive data patterns; Public access block; Data sharing controls', 'Inspection'),
            ('SUB', 'App10-B12(b)', 'CLD-19.2', 'DLP monitoring and response',
             '1. Verify DLP alerts monitored.\n2. Check DLP incidents investigated.\n3. Confirm DLP false positives tuned.\n4. Verify DLP reporting to management.',
             'DLP alert monitoring; Incident investigation records; Tuning records; Management reports', 'Inspection'),
         ]),
        ('B13 — SOC',
         'SOC operations must extend to cloud environments with cloud-specific capabilities, tooling, and skills.',
         [
            ('DOMAIN', 'App10-B13', 'CLD-20', 'PLATFORM', 'SOC',
             'SOC cloud coverage, skills, and tooling'),
            ('SUB', 'App10-B13(a)', 'CLD-20.1', 'SOC cloud coverage',
             '1. Verify SOC monitors cloud environments 24/7.\n2. Check cloud log sources integrated into SOC tooling.\n3. Confirm SOC analysts trained on cloud-specific threats.\n4. Verify cloud-specific use cases in SIEM.',
             'SOC cloud coverage documentation; Log source integration; Analyst training records; Cloud use cases', 'Inspection'),
            ('SUB', 'App10-B13(b)', 'CLD-20.2', 'SOC tooling and automation',
             '1. Verify SOC tooling supports cloud-native events.\n2. Check SOAR playbooks include cloud scenarios.\n3. Confirm automated response capabilities for cloud (quarantine, disable access).\n4. Verify SOC metrics cover cloud incidents.',
             'SOC tooling assessment; SOAR playbooks; Automated response config; SOC metrics', 'Inspection'),
         ]),
        ('B14 — Cyber Response/Recovery',
         'Cloud cyber response and recovery requires cloud-specific procedures, tested recovery, and coordination with CSPs.',
         [
            ('DOMAIN', 'App10-B14', 'CLD-21', 'PLATFORM', 'Cyber Response/Recovery',
             'Cloud-specific cyber response, recovery procedures, and CSP coordination'),
            ('SUB', 'App10-B14(a)', 'CLD-21.1', 'Cloud cyber response plan',
             '1. Verify cyber response plan covers cloud environments.\n2. Check cloud-specific scenarios (ransomware, data exfiltration, account takeover).\n3. Confirm response plan tested.',
             'Cloud cyber response plan; Scenario coverage; Test records', 'Inspection'),
            ('SUB', 'App10-B14(b)', 'CLD-21.2', 'CSP coordination for incidents',
             '1. Verify CSP incident coordination procedures documented.\n2. Check CSP security contact information current.\n3. Confirm CSP notification SLAs understood.\n4. Verify joint incident response tested.',
             'CSP coordination procedures; Contact information; Notification SLAs; Joint test records', 'Inspection'),
            ('SUB', 'App10-B14(c)', 'CLD-21.3', 'Cloud evidence preservation',
             '1. Verify evidence preservation procedures for cloud (snapshots, log exports).\n2. Check evidence collection chain of custody.\n3. Confirm evidence retained for required period.\n4. Verify forensic readiness.',
             'Evidence preservation procedures; Chain of custody; Retention policy; Forensic readiness', 'Inspection'),
            ('SUB', 'App10-B14(d)', 'CLD-21.4', 'Recovery procedures',
             '1. Verify recovery procedures for cloud services.\n2. Check recovery includes security validation before resuming services.\n3. Confirm recovery tested.\n4. Verify recovery communication plan.',
             'Recovery procedures; Security validation checklist; Recovery test results; Communication plan', 'Inspection'),
            ('SUB', 'App10-B14(e)', 'CLD-21.5', 'Post-incident review',
             '1. Verify post-incident review conducted for cloud incidents.\n2. Check lessons learned documented.\n3. Confirm improvements implemented.\n4. Verify management reporting on incidents.',
             'Post-incident review reports; Lessons learned; Improvement tracker; Management reports', 'Inspection'),
            ('SUB', 'App10-B14(f)', 'CLD-21.6', 'Regulatory notification',
             '1. Verify regulatory notification procedures for cloud incidents.\n2. Check notification timelines comply with BNM requirements.\n3. Confirm notification process tested.\n4. Verify notification templates prepared.',
             'Regulatory notification procedure; Timeline compliance; Test records; Notification templates', 'Inspection'),
            ('SUB', 'App10-B14(g)', 'CLD-21.7', 'Cyber insurance coverage',
             '1. Verify cyber insurance covers cloud-related incidents.\n2. Check coverage adequacy relative to cloud exposure.\n3. Confirm policy terms understood.\n4. Verify claims process documented.',
             'Cyber insurance policy; Coverage assessment; Claims process documentation', 'Inspection'),
         ]),
    ]

    write_assessment_rows(ws, r, part_b)


def create_cloud_workbook():
    scope_desc = (
        'This assessment is conducted pursuant to BNM RMiT (BNM/RH/PD 028-98, 28 Nov 2025), Appendix 7 and Appendix 10. '
        'Where the cloud deployment involves emerging technology (AI/ML, DLT, IoT, etc.), the IESP-EmergingTech-WorkProgram.xlsx (Appendix 9) must also be used in conjunction with this workbook.'
    )
    scoping_note = 'Note: Where the cloud deployment involves emerging technology (AI/ML, DLT, IoT, etc.), the IESP-EmergingTech-WorkProgram.xlsx (Appendix 9) must also be used in conjunction with this workbook.'
    wb = build_workbook(
        'BNM RMiT IESP — Cloud Services Assessment',
        'Paragraph 17.1 / Appendix 7, mapped to Appendix 10',
        scope_desc,
        'Assessment covers: (1) Appendix 10 Part A — Cloud Governance (7 areas), (2) Appendix 10 Part B — Cloud Design and Control (14 areas), (3) Appendix 7 Part D — Minimum Controls.',
        'Cloud IESP Assessment (Pre-Implementation or Independent Attestation)',
        add_cloud_sheets,
        scoping_note=scoping_note)
    wb.save('audit-work-programs/IESP-Cloud-WorkProgram.xlsx')
    print('Created IESP-Cloud-WorkProgram.xlsx')


# ═══════════════════════════════════════════════════════════════════
# EMERGING TECH AWP — APPENDIX 9
# ═══════════════════════════════════════════════════════════════════

def add_emerging_tech_sheets(wb):
    ws = wb.create_sheet(title='Appendix 9 - Assessment')
    apply_col_widths(ws)
    r = 1
    write_header_row(ws, r)
    r += 1

    sections = [
        ('BNM Appendix 9 Item 1 — TRMF Governance for Emerging Tech',
         'BNM requires the FI to integrate emerging technology governance into the TRMF. The board must approve the technology risk appetite, and the framework must cover risk assessment, acceptance criteria, and ongoing monitoring. This is a regulatory requirement with 4 sub-items.',
         [
            ('DOMAIN', 'App9-1', 'AI-01', 'ORG', 'TRMF Governance for Emerging Tech',
             'Board-approved governance framework integrated with TRMF for emerging technology'),
            ('SUB', 'App9-1(a)', 'AI-01.1', 'Governance framework and accountability',
             '1. Obtain the FI\'s AI/emerging tech governance framework.\n2. Verify it defines accountability, decision rights, risk management, and ethical principles.\n3. Check designated committee has oversight.\n4. Verify framework approved at board/senior management level.',
             'AI governance framework; Committee ToR; Board/senior management approval records', 'Inspection'),
            ('SUB', 'App9-1(b)', 'AI-01.2', 'Roles and responsibilities',
             '1. Review organisational structure for AI/emerging tech.\n2. Verify defined roles: model owner, developer, validator, risk manager, compliance officer.\n3. Check each deployed model/technology has an accountable owner.\n4. Verify sufficient expertise exists.',
             'Org chart (AI function); Role descriptions; Model ownership register; Skills inventory', 'Inspection / Inquiry'),
            ('SUB', 'App9-1(c)', 'AI-01.3', 'Risk assessment methodology',
             '1. Obtain risk assessment methodology for AI/emerging tech.\n2. Verify it covers: model risk, data risk, bias/fairness, explainability, operational, regulatory.\n3. Check methodology is proportionate to risk tier.\n4. Verify methodology reviewed and updated periodically.',
             'Risk assessment methodology; Risk tier definitions; Review history', 'Inspection'),
            ('SUB', 'App9-1(d)', 'AI-01.4', 'Technology radar and innovation governance',
             '1. Check whether the FI maintains a technology radar or equivalent mechanism.\n2. Verify it tracks emerging technologies and assesses maturity and applicability.\n3. Review the innovation governance process (sandbox/PoC governance).\n4. Check formal evaluation and approval process exists for new technology initiatives.',
             'Technology radar; Innovation governance policy; PoC/sandbox evaluation records; Technology evaluation board minutes', 'Inspection'),
         ]),
        ('BNM Appendix 9 Item 2 — Minimum Production Requirements',
         'BNM requires minimum standards before emerging technology is deployed to production. These include acceptance criteria, security assessment, testing, monitoring, and human oversight. This is a regulatory requirement with 5 sub-items.',
         [
            ('DOMAIN', 'App9-2', 'AI-02', 'WORKLOAD', 'Minimum Production Requirements',
             'Acceptance criteria, security, testing, monitoring, and human oversight for production deployment'),
            ('SUB', 'App9-2(a)', 'AI-02.1', 'Production acceptance criteria',
             '1. Obtain acceptance criteria framework.\n2. Verify criteria cover: performance thresholds, validation results, security assessment, ethical review, operational readiness.\n3. For each in-scope deployment, verify criteria met before production.\n4. Check acceptance formally documented and signed off.',
             'Acceptance criteria framework; Completed acceptance checklists; Sign-off records', 'Inspection'),
            ('SUB', 'App9-2(b)', 'AI-02.2', 'Security assessment for production',
             '1. Verify security assessment conducted before production deployment.\n2. Check assessment covers: access controls, data protection, model integrity, adversarial threats.\n3. Confirm findings remediated before go-live.\n4. Verify security testing includes AI-specific attack vectors.',
             'Security assessment report; Finding remediation evidence; AI-specific security test results', 'Inspection'),
            ('SUB', 'App9-2(c)', 'AI-02.3', 'Pre-production testing',
             '1. Review testing strategy.\n2. Verify: unit, integration, performance, security, UAT, A/B or shadow testing.\n3. Verify test execution records and results.\n4. Confirm exit criteria met before release.',
             'Testing strategy; Test plans and records; Test results; Production release approval', 'Inspection'),
            ('SUB', 'App9-2(d)', 'AI-02.4', 'Production monitoring',
             '1. Identify all models in production.\n2. Verify each has defined performance metrics.\n3. Check performance monitored continuously or at defined intervals.\n4. Verify degradation triggers alerts and investigation.',
             'Production model inventory; Performance metrics per model; Monitoring config; Alert config', 'Inspection'),
            ('SUB', 'App9-2(e)', 'AI-02.5', 'Human oversight and kill-switch',
             '1. Verify human oversight levels defined and proportionate to risk.\n2. Check each production AI system has kill-switch capability.\n3. Verify activation within defined timeframes.\n4. Confirm fallback/manual processes exist.\n5. Verify decision authority defined.',
             'Human oversight framework; Kill-switch documentation; Suspension test results; Fallback process docs', 'Inspection'),
         ]),
        ('Best Practice — Model Governance',
         'Best practice domain beyond BNM requirements. Covers model validation, lifecycle management, and model risk management. Industry frameworks (SR 11-7, SS2/21) inform these controls.',
         [
            ('DOMAIN', '', 'AI-03', 'ORG', 'Model Governance (Best Practice)',
             'Model validation, lifecycle management, and model risk management'),
            ('SUB', '', 'AI-03.1', 'Model inventory and classification',
             '1. Obtain model inventory.\n2. Verify models classified by risk tier (Tier 1/2/3).\n3. Check model owners assigned.\n4. Verify inventory includes model purpose, inputs, outputs, dependencies.',
             'Model inventory; Risk tier classification; Model ownership register', 'Inspection'),
            ('SUB', '', 'AI-03.2', 'Independent model validation',
             '1. Check validation performed independently of development.\n2. Review methodology: conceptual soundness, data adequacy, benchmarking, sensitivity.\n3. Verify validation before deployment and periodically.\n4. Check findings tracked and remediated.',
             'Model validation policy; Independence evidence; Validation reports; Finding tracker', 'Inspection'),
            ('SUB', '', 'AI-03.3', 'Model retraining and recalibration',
             '1. Verify retraining/recalibration procedures exist.\n2. Check triggers for retraining defined (performance drift, data change).\n3. Confirm retraining follows change management.\n4. Verify retrained models re-validated.',
             'Retraining procedures; Trigger definitions; Change records; Re-validation evidence', 'Inspection'),
         ]),
        ('Best Practice — Data Governance',
         'Best practice domain beyond BNM requirements. Covers data quality, lineage, privacy, and representativeness for AI/ML training and inference data.',
         [
            ('DOMAIN', '', 'AI-04', 'WORKLOAD', 'Data Governance (Best Practice)',
             'Data quality, lineage, privacy, and representativeness for AI/ML'),
            ('SUB', '', 'AI-04.1', 'Data quality controls',
             '1. Review data governance framework for AI training/inference data.\n2. Verify data quality controls: completeness, accuracy, timeliness, consistency.\n3. Check data quality metrics defined and monitored.\n4. Verify data quality issues trigger remediation.',
             'Data governance framework; Data quality procedures and metrics; Data quality monitoring', 'Inspection'),
            ('SUB', '', 'AI-04.2', 'Data lineage and provenance',
             '1. Verify data lineage tracked from source to model input.\n2. Check data provenance documented.\n3. Confirm data transformations auditable.\n4. Verify lineage maintained through model updates.',
             'Data lineage documentation; Provenance records; Transformation audit trail', 'Inspection'),
            ('SUB', '', 'AI-04.3', 'Privacy and consent',
             '1. Verify privacy requirements met (anonymisation, consent, purpose limitation).\n2. Check PDPA compliance for training data.\n3. Confirm data retention and deletion procedures.\n4. Verify re-identification risk assessed.',
             'Privacy compliance evidence; PDPA assessment; Retention/deletion procedures; Re-identification risk assessment', 'Inspection'),
         ]),
        ('Best Practice — Bias and Fairness',
         'Best practice domain beyond BNM requirements. Covers bias testing, disparate impact analysis, and ongoing fairness monitoring for AI models.',
         [
            ('DOMAIN', '', 'AI-05', 'WORKLOAD', 'Bias and Fairness (Best Practice)',
             'Bias testing, fairness metrics, and ongoing monitoring for disparate impact'),
            ('SUB', '', 'AI-05.1', 'Bias testing methodology',
             '1. Identify AI models making/supporting customer decisions.\n2. Verify bias testing methodology defined.\n3. Check protected attributes assessed for disparate impact.\n4. Verify bias testing performed before deployment.',
             'Bias testing methodology; Protected attribute list; Pre-deployment bias test results', 'Inspection'),
            ('SUB', '', 'AI-05.2', 'Ongoing fairness monitoring',
             '1. Verify ongoing monitoring for bias drift.\n2. Check fairness metrics defined and tracked.\n3. Confirm bias drift triggers investigation and remediation.\n4. Verify monitoring covers all customer-facing models.',
             'Fairness monitoring config; Fairness metrics; Drift detection; Remediation records', 'Inspection'),
         ]),
        ('Best Practice — Explainability and Transparency',
         'Best practice domain beyond BNM requirements. Covers model explainability, customer disclosure, and decision transparency.',
         [
            ('DOMAIN', '', 'AI-06', 'WORKLOAD', 'Explainability and Transparency (Best Practice)',
             'Model explainability, customer disclosure, and decision interpretability'),
            ('SUB', '', 'AI-06.1', 'Model explainability',
             '1. Verify explainability methods implemented (SHAP, LIME, feature importance).\n2. Check explanations can be provided for individual decisions.\n3. Confirm explainability proportionate to risk and customer impact.\n4. Verify explainability tested.',
             'Explainability methods; Individual decision explanations; Explainability test results', 'Inspection'),
            ('SUB', '', 'AI-06.2', 'Customer disclosure',
             '1. Identify AI systems interacting with customers.\n2. Verify customers informed of AI use in decision-making.\n3. Check complaints/appeals process exists.\n4. Review sample disclosures.',
             'Disclosure policy; Customer notification evidence; Complaints process; Sample disclosures', 'Inspection'),
         ]),
        ('Best Practice — Human Oversight',
         'Best practice domain beyond BNM requirements. Covers human-in-the-loop controls, escalation, and override capabilities.',
         [
            ('DOMAIN', '', 'AI-07', 'WORKLOAD', 'Human Oversight (Best Practice)',
             'Human-in-the-loop controls, override capability, and escalation'),
            ('SUB', '', 'AI-07.1', 'Human-in-the-loop design',
             '1. Verify human oversight levels defined per model risk tier.\n2. Check automated decisions can be overridden by humans.\n3. Verify escalation procedures for anomalous model behaviour.\n4. Confirm override decisions logged and auditable.',
             'Human oversight framework; Override capability; Escalation procedures; Override audit trail', 'Inspection'),
            ('SUB', '', 'AI-07.2', 'AI incident management',
             '1. Review IR procedure for AI incidents (model failure, biased outcomes, data poisoning, adversarial attacks).\n2. Verify AI incident categories and severity levels defined.\n3. Check incident log for past 12 months.\n4. Verify lessons learned fed back into governance.',
             'AI incident management procedure; AI incident categories; AI incident log; Lessons learned register', 'Inspection'),
            ('SUB', '', 'AI-07.3', 'Outcome monitoring for unintended consequences',
             '1. Verify outcome tracking implemented for AI decisions.\n2. Check monitoring for unintended consequences.\n3. Confirm anomalous outcomes trigger investigation.\n4. Verify feedback loop to model improvement.',
             'Outcome monitoring config; Unintended consequence tracking; Investigation records; Feedback loop evidence', 'Inspection'),
         ]),
    ]

    write_assessment_rows(ws, r, sections)


def create_emerging_tech_workbook():
    scope_desc = (
        'This assessment is conducted pursuant to BNM RMiT (BNM/RH/PD 028-98, 28 Nov 2025), Appendix 7 and Appendix 9. '
        'Where emerging technology is deployed on cloud infrastructure, the IESP-Cloud-WorkProgram.xlsx (Appendix 10) must also be used in conjunction with this workbook.'
    )
    scoping_note = 'Note: Where emerging technology is deployed on cloud infrastructure, the IESP-Cloud-WorkProgram.xlsx (Appendix 10) must also be used in conjunction with this workbook.'
    wb = build_workbook(
        'BNM RMiT IESP — Emerging Technology Assessment',
        'Paragraph 17.1 / Appendix 7, mapped to Appendix 9',
        scope_desc,
        'Assessment covers: (1) Appendix 9 — Emerging Technology (2 BNM items + 5 best practice domains), (2) Appendix 7 Part D — Minimum Controls.',
        'Emerging Technology IESP Assessment (Pre-Implementation or Independent Attestation)',
        add_emerging_tech_sheets,
        scoping_note=scoping_note)
    wb.save('audit-work-programs/IESP-EmergingTech-WorkProgram.xlsx')
    print('Created IESP-EmergingTech-WorkProgram.xlsx')


# ═══════════════════════════════════════════════════════════════════
# DIGITAL SERVICES AWP — 16.4/16.5 + Part D
# ═══════════════════════════════════════════════════════════════════

def add_digital_services_sheets(wb):
    ws = wb.create_sheet(title='Digital Services Assessment')
    apply_col_widths(ws)
    r = 1
    write_header_row(ws, r)
    r += 1

    sections = [
        ('Domain 1 — Access Control (Part D 1a + 2a-c)',
         'Customer-facing authentication is a primary defence against fraud and account takeover. Expect MFA, secure session management, brute-force protection. Weak authentication directly impacts customers.',
         [
            ('DOMAIN', 'App7-D1(a)/D2(a)', 'DS-01', 'WORKLOAD', 'Access Control',
             'Customer authentication, identity verification, session management, and brute-force protection'),
            ('SUB', '16.4/16.5', 'DS-01.1', 'Customer authentication',
             '1. Review authentication mechanisms (password, OTP, biometric, device binding).\n2. Verify MFA enforced for login and high-risk transactions.\n3. Check credential policy: complexity, history, expiry, lockout.\n4. Verify session tokens securely managed.',
             'Authentication design; MFA config; Credential policy; Session management config', 'Inspection'),
            ('SUB', '16.4/16.5', 'DS-01.2', 'Digital identity verification (eKYC)',
             '1. Review eKYC process.\n2. Verify identity document validation (OCR, NFC, database).\n3. Check liveness detection.\n4. Verify results logged and auditable.\n5. Assess fraud detection controls during onboarding.',
             'eKYC process flow; Identity verification config; Liveness detection; Audit logs; Fraud detection rules', 'Inspection'),
            ('SUB', '16.4/16.5', 'DS-01.3', 'Authorisation and RBAC',
             '1. Review authorisation model (RBAC, ABAC).\n2. Verify access based on role/entitlements.\n3. Check privilege escalation prevented.\n4. Verify API authorisation.\n5. Test for IDOR.',
             'Authorisation model; Role-to-function mapping; API auth config; Pen test results (IDOR)', 'Inspection'),
            ('SUB', '16.4/16.5', 'DS-01.4', 'Session management',
             '1. Review session management.\n2. Verify tokens cryptographically random.\n3. Check idle and absolute timeouts.\n4. Verify session invalidated on logout/password change.\n5. Check concurrent session controls.',
             'Session management config; Token generation; Timeout settings; Concurrent session policy', 'Inspection'),
            ('SUB', '16.4/16.5', 'DS-01.5', 'Account lockout and brute-force protection',
             '1. Review lockout policy.\n2. Verify rate limiting.\n3. Check CAPTCHA/bot protection.\n4. Verify lockout notifications.\n5. Test bypass resistance.',
             'Lockout config; Rate limiting config; CAPTCHA/bot protection; Notification config; Bypass test results', 'Inspection'),
         ]),
        ('Domain 2 — Online Transaction Security (Part D 2)',
         'Transaction integrity and fraud monitoring protect customer funds. Expect step-up authentication, transaction binding, fraud monitoring rules, encryption. Transaction fraud is a direct financial and reputational impact.',
         [
            ('DOMAIN', 'App7-D2(b)', 'DS-02', 'WORKLOAD', 'Transaction Security',
             'Transaction authentication, limits, fraud monitoring, encryption, and non-repudiation'),
            ('SUB', '16.4/16.5', 'DS-02.1', 'Transaction authentication',
             '1. Review transaction authentication (signing, OTP, biometric, TAC).\n2. Verify step-up for high-risk transactions.\n3. Check confirmation presented before execution.\n4. Verify authentication bound to specific transaction.\n5. Assess MITM/MITB protections.',
             'Transaction auth design; Step-up config; Confirmation flow; Transaction binding; MitM protections', 'Inspection'),
            ('SUB', '16.4/16.5', 'DS-02.2', 'Transaction limits and fraud monitoring',
             '1. Obtain transaction limits.\n2. Verify enforced at application and backend.\n3. Review fraud monitoring rules.\n4. Check suspicious transactions trigger alerts/blocks.\n5. Verify customer limit management.',
             'Transaction limit config; Fraud monitoring rules; Sample fraud alerts; Customer limit management', 'Inspection'),
            ('SUB', '16.4/16.5', 'DS-02.3', 'End-to-end encryption',
             '1. Verify TLS 1.2+ enforced.\n2. Check certificate config (valid, HSTS).\n3. Verify no sensitive data in URLs/logs.\n4. Check certificate pinning for mobile.\n5. Verify mTLS for B2B.',
             'TLS config (SSL Labs); Certificate details; HSTS config; Log review; Cert pinning config; mTLS config', 'Inspection'),
            ('SUB', '16.4/16.5', 'DS-02.4', 'Secure API design',
             '1. Review API design against OWASP API Top 10.\n2. Verify API auth (OAuth 2.0, API keys, mTLS).\n3. Check rate limiting.\n4. Verify input validation and output encoding.\n5. Check API versioning.\n6. Review API gateway config.',
             'API documentation; OWASP assessment; API auth config; Rate limiting; API gateway config', 'Inspection'),
            ('SUB', '16.4/16.5', 'DS-02.5', 'Non-repudiation',
             '1. Identify critical transaction types.\n2. Verify records capture who, what, when, where, how authenticated.\n3. Check records tamper-proof.\n4. Verify audit trail retention.',
             'Non-repudiation design; Transaction record format; Tamper-proofing mechanism; Retention config', 'Inspection'),
         ]),
        ('Domain 3 — Mobile Device Security (Part D 2e)',
         'Mobile applications extend the attack surface to customer devices. Expect root detection, secure storage, RASP, certificate pinning. Mobile apps are exposed to reverse engineering and tampering.',
         [
            ('DOMAIN', 'App7-D2(e)', 'DS-03', 'WORKLOAD', 'Mobile Security',
             'Mobile app security controls, data protection, distribution, and threat mitigations'),
            ('SUB', '16.4/16.5', 'DS-03.1', 'Mobile app security controls',
             '1. Review mobile security architecture.\n2. Verify root/jailbreak detection.\n3. Check no plaintext sensitive data on device.\n4. Verify code obfuscation and anti-tampering.\n5. Check no data leakage in logs/screenshots/clipboard.',
             'Mobile security architecture; Root/jailbreak detection; Data storage review; Obfuscation evidence; Leakage assessment', 'Inspection'),
            ('SUB', '16.4/16.5', 'DS-03.2', 'App distribution and updates',
             '1. Verify official app store only.\n2. Check forced updates for critical patches.\n3. Verify app signing and integrity.\n4. Check push notifications don\'t include sensitive data.\n5. Review MDM/MAM requirements.',
             'App store distribution; Minimum version enforcement; App signing config; Push notification policy; MDM/MAM policy', 'Inspection'),
            ('SUB', '16.4/16.5', 'DS-03.3', 'Mobile threat protections',
             '1. Check for RASP.\n2. Verify anti-debugging, emulator, hooking framework detection.\n3. Check proxy/MitM detection.\n4. Verify device binding.',
             'RASP config; Anti-debugging/emulator detection; Proxy detection; Device binding mechanism', 'Inspection'),
         ]),
        ('Domain 4 — Data Integrity (Part D 2d)',
         'Input validation and anti-replay controls prevent data manipulation. Expect server-side validation, injection testing, integrity controls. Data integrity failures can lead to financial loss.',
         [
            ('DOMAIN', 'App7-D2(d)', 'DS-04', 'WORKLOAD', 'Data Integrity',
             'Input validation, processing integrity, and anti-replay/anti-tampering controls'),
            ('SUB', '16.4/16.5', 'DS-04.1', 'Input validation',
             '1. Review server-side input validation.\n2. Verify: data type, length, range, format validation.\n3. Test for injection attacks (SQL, XSS, command, LDAP).\n4. Verify output encoding.\n5. Check file upload validation.',
             'Input validation design; Server-side validation config; Pen test results (injection); Output encoding; File upload config', 'Inspection'),
            ('SUB', '16.4/16.5', 'DS-04.2', 'Processing and storage integrity',
             '1. Review integrity controls in processing chain.\n2. Verify checksums/hash verification between components.\n3. Check database integrity controls.\n4. Verify reconciliation with core systems.\n5. Confirm integrity violation alerts.',
             'Integrity control design; Checksum implementation; Database integrity config; Reconciliation procedures; Alert config', 'Inspection'),
            ('SUB', '16.4/16.5', 'DS-04.3', 'Replay and manipulation protection',
             '1. Verify anti-replay mechanisms (nonce, timestamp, sequence).\n2. Check transaction integrity in transit.\n3. Verify duplicate detection.\n4. Test for parameter tampering.',
             'Anti-replay mechanism; Transit integrity protection; Duplicate detection config; Parameter tampering test results', 'Inspection'),
         ]),
        ('Supplementary — Application Security Testing',
         'Comprehensive security testing before launch is the last gate. Expect SAST, DAST, pen testing, SSDLC, privacy assessment. Launching without testing exposes customers to known vulnerabilities.',
         [
            ('DOMAIN', '', 'DS-05', 'WORKLOAD', 'Application Security Testing',
             'SAST, DAST, pen testing, SSDLC, privacy, and availability'),
            ('SUB', '', 'DS-05.1', 'Comprehensive security testing',
             '1. Obtain testing scope and plan.\n2. Verify: SAST, DAST, pen testing, API security testing, infrastructure VA.\n3. Review results and severity distribution.\n4. Verify high/critical findings remediated before launch.\n5. Confirm qualified testers.',
             'Security testing plan; SAST/DAST/pen test/API/VA reports; Remediation evidence; Tester qualifications', 'Inspection'),
            ('SUB', '', 'DS-05.2', 'Secure SDLC',
             '1. Review SSDLC.\n2. Verify security requirements at design phase.\n3. Check threat modelling performed.\n4. Verify secure coding standards enforced.\n5. Confirm security sign-off before release.',
             'SSDLC documentation; Security requirements; Threat model; Secure coding standards; Security sign-off records', 'Inspection'),
            ('SUB', '', 'DS-05.3', 'Privacy and data protection',
             '1. Review PIA for the digital service.\n2. Verify data minimisation.\n3. Check consent mechanisms.\n4. Verify retention/deletion policies.\n5. Confirm privacy notices clear.',
             'PIA; Data collection inventory; Consent mechanism; Retention/deletion config; Privacy notice review', 'Inspection'),
            ('SUB', '', 'DS-05.4', 'Business continuity and availability',
             '1. Review availability targets.\n2. Verify HA architecture.\n3. Check monitoring and alerting.\n4. Verify DR plan covers the service.\n5. Review most recent DR test.',
             'Availability targets; HA architecture; Monitoring config; DR plan; DR test results', 'Inspection'),
         ]),
    ]

    write_assessment_rows(ws, r, sections)


def create_digital_services_workbook():
    wb = build_workbook(
        'BNM RMiT IESP — Digital Services Enhancement Assessment',
        'Paragraphs 16.4, 16.5 / Appendix 7 Part D',
        'This assessment is conducted pursuant to BNM RMiT (BNM/RH/PD 028-98, 28 Nov 2025), Paragraphs 16.4 and 16.5.',
        'Assessment covers: (1) Digital Services Security Assessment (5 domains), (2) Appendix 7 Part D — Minimum Controls.',
        'Digital Services Enhancement IESP Assessment (Pre-Launch)',
        add_digital_services_sheets)
    wb.save('audit-work-programs/IESP-DigitalServices-WorkProgram.xlsx')
    print('Created IESP-DigitalServices-WorkProgram.xlsx')


# ═══════════════════════════════════════════════════════════════════
# DCRA AWP — Clauses 10.24-10.28
# ═══════════════════════════════════════════════════════════════════

def add_dcra_sheets(wb):
    ws = wb.create_sheet(title='DCRA Assessment')
    apply_col_widths(ws)
    r = 1
    write_header_row(ws, r)
    r += 1

    sections = [
        ('Domain 1 — DC Resilience and Availability Objectives (10.24)',
         'DC resilience objectives must align with business recovery requirements. Expect quantified targets (availability %, RTO, RPO) approved by committee. Misaligned objectives mean the DC cannot support business recovery.',
         [
            ('DOMAIN', '10.24', 'DCRA-01', 'ORG', 'DC Resilience Objectives',
             'Resilience targets aligned with business recovery requirements'),
            ('SUB', '10.24', 'DCRA-01.1', 'Objectives alignment and quantified targets',
             '1. Obtain DC resilience strategy.\n2. Extract resilience targets (availability %, RTO, RPO).\n3. Compare against approved business recovery objectives.\n4. Check targets approved by appropriate committee.\n5. Verify periodic review (annual or upon material change).',
             'DC resilience strategy; Business recovery objectives matrix; Committee approval minutes; Review records', 'Inspection'),
         ]),
        ('Domain 2 — Redundancy and Single Points of Failure (10.25)',
         'Critical infrastructure must have no single points of failure. Expect N+1 or 2N redundancy, diverse paths, tested failover. A single failure in power, cooling, or network can cause a DC outage.',
         [
            ('DOMAIN', '10.25', 'DCRA-02', 'PLATFORM', 'Redundancy and SPOF',
             'Redundant capacity, diverse paths, SPOF analysis, and failover testing'),
            ('SUB', '10.25', 'DCRA-02.1', 'Redundant capacity design',
             '1. Obtain redundancy design (N+1, 2N, 2N+1).\n2. For each critical layer (power, cooling, network), confirm redundancy.\n3. Verify actual capacity against design through site inspection.\n4. Check redundant capacity handles full load on failover.',
             'Redundancy design specs; Capacity planning reports; Site inspection observations', 'Inspection / Observation'),
            ('SUB', '', 'DCRA-02.2', 'Multiple distribution paths',
             '1. Review electrical single-line diagrams for dual feeds.\n2. Review network diagrams for diverse carrier paths.\n3. Confirm physical path diversity.\n4. Verify through site walk.',
             'Electrical single-line diagrams; Network path diversity docs; Site walk observations', 'Inspection / Observation'),
            ('SUB', '', 'DCRA-02.3', 'SPOF analysis',
             '1. Obtain SPOF analysis per DC.\n2. Walk through each SPOF and confirm mitigations.\n3. Identify unrecognised SPOFs during independent walkthrough.\n4. Check SPOF analysis periodically refreshed.',
             'SPOF analysis/register; Mitigation action plans; Site walk findings', 'Inspection / Observation'),
            ('SUB', '', 'DCRA-02.4', 'Failover testing',
             '1. Obtain failover test plans and results (UPS, generators, cooling, switches).\n2. Review test frequency (at least annual).\n3. Confirm tests simulate actual failure conditions.\n4. Check test failures remediated.',
             'Failover test plans; Test result reports; Corrective action records', 'Inspection'),
         ]),
        ('Domain 3 — Physical Security and Environmental Controls (10.26)',
         'Physical security protects against unauthorised access and environmental threats. Expect multi-layered access, environmental monitoring, fire suppression, hazard assessment. Physical compromise bypasses all logical controls.',
         [
            ('DOMAIN', '10.26', 'DCRA-03', 'PLATFORM', 'Physical Security & Environmental',
             'Multi-layered physical access, environmental monitoring, fire suppression, and hazard assessment'),
            ('SUB', '10.26', 'DCRA-03.1', 'Dedicated space and physical separation',
             '1. Confirm DC in dedicated space.\n2. Verify physical separation from public areas.\n3. Check walls extend true floor to true ceiling.',
             'Site inspection observations; Building layout plans; Lease/facility agreements', 'Observation'),
            ('SUB', '10.26', 'DCRA-03.2', 'Multi-layered access controls',
             '1. Walk access path from perimeter to rack.\n2. Verify independent controls at each layer.\n3. Test access mechanisms.\n4. Review provisioning/de-provisioning.\n5. Sample-check access list against HR for terminated staff.',
             'Physical security policy; Access control config; Access list vs HR reconciliation; Site walk observations', 'Inspection / Observation'),
            ('SUB', '10.26', 'DCRA-03.3', 'Disaster-prone location assessment',
             '1. Obtain site risk assessment (flood, earthquake, lightning).\n2. Cross-reference against hazard maps.\n3. Verify mitigations.\n4. Check assessment periodically updated.',
             'Site risk assessment; Hazard map cross-reference; Mitigation evidence', 'Inspection'),
            ('SUB', '10.26', 'DCRA-03.4', 'Electrical infrastructure',
             '1. Review UPS configuration (N+1 or 2N).\n2. Verify UPS runtime (min 15 min).\n3. Confirm generator auto-start and fuel autonomy (48-72 hrs).\n4. Check ATS operation and tests.\n5. Verify fuel replenishment contracts.',
             'UPS specs and test reports; Generator load test reports; ATS test records; Fuel supply contracts', 'Inspection / Observation'),
            ('SUB', '10.26', 'DCRA-03.5', 'Thermal management',
             '1. Review cooling design (N+1 min).\n2. Check cooling load vs capacity.\n3. Verify rack-level monitoring.\n4. Confirm alerting thresholds.\n5. Review hot/cold aisle containment.',
             'Cooling design docs; Capacity reports; Monitoring config; Alert thresholds', 'Inspection / Observation'),
            ('SUB', '10.26', 'DCRA-03.6', 'Fire suppression and monitoring',
             '1. Confirm VESDA or equivalent.\n2. Verify gas-based suppression with current inspections.\n3. Check water leak detection.\n4. Verify 24/7 monitoring and alerting.\n5. Review environmental incident response procedures.',
             'Fire suppression certificates; VESDA config; Water leak detection; Monitoring dashboard; Environmental IR procedures', 'Inspection / Observation'),
            ('SUB', '', 'DCRA-03.7', 'Hardware lifecycle management',
             '1. Obtain hardware asset register.\n2. Sample-check physical assets against register.\n3. Verify EOL/EOS hardware tracked with replacement plan.\n4. Check secure data destruction for decommissioned equipment.',
             'Hardware asset register; Physical verification; EOL/EOS tracker; Data destruction certificates', 'Inspection / Observation'),
         ]),
        ('Domain 4 — DC Operations and Control Procedures (10.27)',
         'Operational controls ensure stable, monitored DC operations. Expect automated monitoring, controlled changes, incident management with RCA. Uncontrolled operations lead to avoidable outages.',
         [
            ('DOMAIN', '10.27', 'DCRA-04', 'PLATFORM', 'DC Operations',
             'Automated monitoring, batch processing, change management, and incident handling'),
            ('SUB', '10.27', 'DCRA-04.1', 'Automated monitoring',
             '1. Obtain monitoring tool inventory (DCIM, BMS, network).\n2. Verify coverage of critical parameters.\n3. Test alerts for threshold breaches.\n4. Confirm NOC/SOC integration.',
             'Tool inventory; Monitoring coverage matrix; Sample alerts; NOC/SOC integration', 'Inspection'),
            ('SUB', '10.27', 'DCRA-04.2', 'Batch processing controls',
             '1. Obtain batch job procedures.\n2. Verify jobs have defined run windows, dependencies, error handling.\n3. Review failure logs (3 months).\n4. Check schedule changes controlled.',
             'Batch scheduling procedures; Job scheduler config; Failure logs; Change approval records', 'Inspection'),
            ('SUB', '10.27', 'DCRA-04.3', 'Change management',
             '1. Obtain DC change management procedure.\n2. Sample 10 infrastructure changes.\n3. Verify full lifecycle (request through post-impl review).\n4. Check emergency changes follow defined process.',
             'Change management procedure; 10 sampled change records; Emergency change records', 'Inspection'),
            ('SUB', '10.27', 'DCRA-04.4', 'Error and incident handling',
             '1. Obtain DC incident management procedure.\n2. Review incident log (6 months).\n3. Verify RCA for significant incidents.\n4. Check RCA findings led to preventive actions.',
             'Incident management procedure; Incident log (6 months); RCA reports; Preventive action tracker', 'Inspection'),
         ]),
        ('Domain 5 — Segregation of Incompatible Activities (10.28)',
         'Segregation prevents unauthorised changes and maintains environment integrity. Expect production isolation, controlled vendor access, SoD enforcement. Without segregation, a single actor can compromise production.',
         [
            ('DOMAIN', '10.28', 'DCRA-05', 'PLATFORM', 'Segregation',
             'Vendor access controls, production segregation, and segregation of duties'),
            ('SUB', '10.28', 'DCRA-05.1', 'Vendor access controls',
             '1. Review vendor/third-party access policy.\n2. Check pre-approved access requests.\n3. Verify escort requirements.\n4. Sample-check vendor access logs.\n5. Confirm access revoked promptly.',
             'Vendor access policy; Access request forms; Escort logs; Access log reconciliation', 'Inspection'),
            ('SUB', '', 'DCRA-05.2', 'Production segregation',
             '1. Review logical and physical segregation (prod vs non-prod).\n2. Verify developers have no direct production access.\n3. Check network segmentation.\n4. Confirm non-prod data masked.',
             'Environment segregation architecture; Access control matrices; Network segmentation; Data masking procedures', 'Inspection'),
            ('SUB', '', 'DCRA-05.3', 'Segregation of duties',
             '1. Obtain DC operations RACI.\n2. Verify incompatible duties separated.\n3. Check privileged access limited and monitored.\n4. Review privileged access usage logs (3 months).',
             'RACI matrix; Role definitions; Privileged access list; Privileged access usage logs', 'Inspection'),
         ]),
    ]

    write_assessment_rows(ws, r, sections)


def create_dcra_workbook():
    wb = build_workbook(
        'BNM RMiT IESP — Data Centre Resilience Assessment (DCRA)',
        'Paragraph 14.1, mapped to paragraphs 10.24 to 10.28',
        'This assessment is conducted pursuant to BNM RMiT (BNM/RH/PD 028-98, 28 Nov 2025), Paragraph 14.1.',
        'Assessment covers: (1) DCRA — DC resilience, redundancy, physical security, operations, segregation (10.24-10.28, 5 domains), (2) Appendix 7 Part D — Minimum Controls.',
        'Data Centre Resilience Assessment (DCRA)',
        add_dcra_sheets)
    wb.save('audit-work-programs/IESP-DCRA-WorkProgram.xlsx')
    print('Created IESP-DCRA-WorkProgram.xlsx')


# ═══════════════════════════════════════════════════════════════════
# NRA AWP — Clauses 10.36-10.43
# ═══════════════════════════════════════════════════════════════════

def add_nra_sheets(wb):
    ws = wb.create_sheet(title='NRA Assessment')
    apply_col_widths(ws)
    r = 1
    write_header_row(ws, r)
    r += 1

    sections = [
        ('Domain 1 — Network Design (10.36)',
         'Network must be designed for current and future capacity. Expect documented capacity planning, scalability, performance SLAs. Under-capacity networks degrade service availability.',
         [
            ('DOMAIN', '10.36', 'NRA-01', 'ORG', 'Network Design',
             'Capacity planning, scalability, and performance SLAs'),
            ('SUB', '10.36', 'NRA-01.1', 'Capacity requirements and planning',
             '1. Obtain capacity planning document.\n2. Review bandwidth utilisation across critical links.\n3. Verify 2-year growth projections.\n4. Check utilisation thresholds trigger augmentation.',
             'Capacity planning document; Bandwidth utilisation reports (12 months); Growth projection model', 'Inspection'),
            ('SUB', '10.36', 'NRA-01.2', 'Scalability and technology currency',
             '1. Review architecture for modular/scalable design.\n2. Assess whether it can accommodate growth without redesign.\n3. Check technology currency — platforms within vendor support.',
             'Architecture design principles; Technology lifecycle tracker; Scalability assessment', 'Inspection'),
            ('SUB', '', 'NRA-01.3', 'Performance SLAs',
             '1. Obtain SLA definitions (latency, jitter, packet loss, availability).\n2. Verify monitoring tools.\n3. Review SLA performance (6 months).\n4. Check breaches trigger investigation.',
             'SLA definitions; Performance monitoring reports; SLA breach incident records', 'Inspection'),
         ]),
        ('Domain 2 — Network Resilience (10.37)',
         'Network redundancy prevents outages. Expect diverse paths, HA pairs, carrier diversity. Network outage impacts all services.',
         [
            ('DOMAIN', '10.37', 'NRA-02', 'PLATFORM', 'Network Resilience',
             'Redundancy, diverse paths, HA pairs, and carrier diversity'),
            ('SUB', '10.37', 'NRA-02.1', 'Redundancy for critical paths',
             '1. Review topology for redundancy at each layer.\n2. Verify redundant paths (dual WAN, diverse carriers).\n3. Confirm redundant devices (HA pairs).\n4. Verify physical route diversity.',
             'Network topology with redundancy; Carrier diversity docs; HA configuration evidence', 'Inspection'),
            ('SUB', '10.37', 'NRA-02.2', 'Carrier/ISP management',
             '1. Review carrier contracts and SLAs.\n2. Verify at least two independent carriers.\n3. Check carrier performance against SLAs (12 months).\n4. Confirm escalation procedures.',
             'Carrier contracts; Carrier diversity mapping; Performance reports; Escalation contacts', 'Inspection'),
         ]),
        ('Domain 3 — Resilience Measures (10.38)',
         'Tested failover and DR ensure continuity. Expect regular failover tests and network DR plans.',
         [
            ('DOMAIN', '10.38', 'NRA-03', 'PLATFORM', 'Resilience Measures',
             'Failover testing and network DR planning'),
            ('SUB', '10.38', 'NRA-03.1', 'Failover testing',
             '1. Obtain failover test plan.\n2. Review most recent tests (WAN, internet, core switches, firewalls).\n3. Verify failover within acceptable times.\n4. Check failures remediated.\n5. Confirm realistic failure simulation.',
             'Failover test plans; Test execution reports; Remediation records', 'Inspection'),
            ('SUB', '10.38', 'NRA-03.2', 'Network DR',
             '1. Review network DR plan.\n2. Verify recovery procedures per critical segment.\n3. Check network DR tested in DR exercises.\n4. Confirm network RTO aligns with business requirements.',
             'Network DR plan; DR test results (network); RTO alignment mapping', 'Inspection'),
         ]),
        ('Domain 4 — Network Monitoring (10.39)',
         'Network visibility enables performance management and threat detection. Expect comprehensive monitoring, flow analysis, anomaly detection, trend reporting. Unmonitored networks hide degradation and threats.',
         [
            ('DOMAIN', '10.39', 'NRA-04', 'PLATFORM', 'Network Monitoring',
             'Comprehensive monitoring, traffic analysis, anomaly detection, and trend reporting'),
            ('SUB', '10.39', 'NRA-04.1', 'Comprehensive monitoring',
             '1. Obtain monitoring tool inventory.\n2. Verify all critical devices/links monitored.\n3. Check parameter coverage (availability, performance, errors).\n4. Confirm 24/7 alerting to NOC.',
             'Monitoring tool inventory; Device coverage report; Parameter coverage matrix; NOC alerting config', 'Inspection'),
            ('SUB', '10.39', 'NRA-04.2', 'Traffic analysis and anomaly detection',
             '1. Confirm flow analysis tools deployed.\n2. Review traffic baselines and anomaly thresholds.\n3. Sample 5 anomaly alerts — verify investigation.\n4. Verify SIEM/SOC integration.',
             'Flow analysis config; Traffic baselines; Sample anomaly investigation; SIEM integration', 'Inspection'),
            ('SUB', '', 'NRA-04.3', 'Performance trend analysis',
             '1. Obtain regular performance reports.\n2. Verify trend analysis (utilisation, latency, availability).\n3. Check trends inform capacity planning.\n4. Confirm management review.',
             'Performance reports; Trend analysis dashboards; Management review records', 'Inspection'),
         ]),
        ('Domain 5 — Confidentiality, Integrity, Availability (10.40)',
         'Network CIA requires layered controls. Expect encryption, NAC, device hardening, configuration integrity monitoring. Network compromise provides access to all connected systems.',
         [
            ('DOMAIN', '10.40', 'NRA-05', 'PLATFORM', 'Confidentiality/Integrity/Availability',
             'Encryption in transit, access controls, device hardening, and configuration integrity'),
            ('SUB', '10.40', 'NRA-05.1', 'Encryption in transit',
             '1. Identify links over untrusted networks.\n2. Verify encryption (IPSec, TLS, MACsec).\n3. Check standards (TLS 1.2+, AES-256).\n4. Verify certificate management.',
             'Encryption standards policy; VPN/TLS config; Certificate inventory', 'Inspection'),
            ('SUB', '10.40', 'NRA-05.2', 'Access controls and authentication',
             '1. Review NAC mechanisms (802.1X).\n2. Verify management access uses RADIUS/TACACS+ with MFA.\n3. Check default credentials changed.\n4. Verify unused ports disabled.\n5. Sample 10 device configs.',
             'NAC config; RADIUS/TACACS+ config; Sample device configs; Port status reports', 'Inspection'),
            ('SUB', '10.40', 'NRA-05.3', 'Configuration integrity',
             '1. Verify daily config backups.\n2. Check drift detection and alerting.\n3. Confirm baseline configs exist per device type.\n4. Verify secure backup storage.',
             'Config backup schedule; Drift monitoring tool; Baseline config library; Backup storage', 'Inspection'),
         ]),
        ('Domain 6 — Network Blueprint (10.41)',
         'A current network blueprint is essential for security and change management. Expect a comprehensive, version-controlled, approved blueprint. Stale blueprints lead to security gaps.',
         [
            ('DOMAIN', '10.41', 'NRA-06', 'ORG', 'Network Blueprint',
             'Comprehensive, version-controlled, and approved network design documentation'),
            ('SUB', '10.41', 'NRA-06.1', 'Current blueprint',
             '1. Obtain network design blueprint.\n2. Verify it covers: logical/physical topology, IP addressing, VLAN design, routing, security zones, external connectivity.\n3. Check it reflects current state.\n4. Verify version-controlled.',
             'Network design blueprint; Version history; Change control records', 'Inspection'),
            ('SUB', '10.41', 'NRA-06.2', 'Review and approval',
             '1. Check blueprint approved by appropriate authority.\n2. Verify reviewed at least annually.\n3. Confirm accessible to relevant teams.',
             'Approval records; Review schedule; Access/distribution records', 'Inspection'),
         ]),
        ('Domain 7 — Network Logging (10.42)',
         'Network device logging supports incident investigation and compliance. Expect centralised, tamper-protected, time-synchronised logs with defined retention. Without logs, incidents cannot be investigated.',
         [
            ('DOMAIN', '10.42', 'NRA-07', 'PLATFORM', 'Network Logging',
             'Comprehensive logging, centralised management, retention, and tamper protection'),
            ('SUB', '10.42', 'NRA-07.1', 'Comprehensive logging',
             '1. Review logging config on 10 sample devices.\n2. Verify security events logged (auth, config changes, ACL, admin actions).\n3. Confirm centralised log management.\n4. Verify NTP synchronisation.',
             'Sample device logging configs; Centralised log system; NTP config', 'Inspection'),
            ('SUB', '10.42', 'NRA-07.2', 'Log retention and review',
             '1. Check retention against policy (min 12 months online).\n2. Verify tamper protection.\n3. Confirm logs reviewed through SIEM or manual process.\n4. Sample 5 events and trace to response.',
             'Log retention policy; Log protection mechanisms; SIEM rules; Sample event investigation records', 'Inspection'),
         ]),
        ('Domain 8 — Network Segmentation (10.43)',
         'Network segmentation limits blast radius and lateral movement. Expect zone-based segmentation, micro-segmentation for critical assets, wireless isolation. Flat networks allow unrestricted lateral movement.',
         [
            ('DOMAIN', '10.43', 'NRA-08', 'PLATFORM', 'Network Segmentation',
             'Zone-based segmentation, micro-segmentation, and wireless network security'),
            ('SUB', '10.43', 'NRA-08.1', 'Segmentation strategy',
             '1. Obtain segmentation policy and design.\n2. Verify: prod vs non-prod, internal vs DMZ, PCI zones, guest/IoT.\n3. Review firewall/ACL rules.\n4. Conduct sample traffic flow tests.',
             'Segmentation policy; Firewall/ACL rule sets; Traffic flow test results', 'Inspection'),
            ('SUB', '10.43', 'NRA-08.2', 'Micro-segmentation for critical assets',
             '1. Identify most critical segments.\n2. Verify additional controls beyond standard segmentation.\n3. Check application-layer segmentation.\n4. Verify explicit authorisation required.',
             'Critical segment identification; Additional segmentation controls; Access authorisation records', 'Inspection'),
            ('SUB', '', 'NRA-08.3', 'Wireless network security',
             '1. Identify all wireless networks.\n2. Verify segmentation from wired corporate.\n3. Check WPA3-Enterprise or WPA2-Enterprise minimum.\n4. Verify rogue AP detection.\n5. Review wireless security assessment.',
             'Wireless inventory; Segmentation design; Auth config; Rogue AP detection; Wireless security reports', 'Inspection'),
         ]),
    ]

    write_assessment_rows(ws, r, sections)


def create_nra_workbook():
    wb = build_workbook(
        'BNM RMiT IESP — Network Resilience Assessment (NRA)',
        'Paragraph 14.2, mapped to paragraphs 10.36 to 10.43',
        'This assessment is conducted pursuant to BNM RMiT (BNM/RH/PD 028-98, 28 Nov 2025), Paragraph 14.2.',
        'Assessment covers: (1) NRA — network design, resilience, monitoring, security, segmentation (10.36-10.43, 8 domains), (2) Appendix 7 Part D — Minimum Controls.',
        'Network Resilience Assessment (NRA)',
        add_nra_sheets)
    wb.save('audit-work-programs/IESP-NRA-WorkProgram.xlsx')
    print('Created IESP-NRA-WorkProgram.xlsx')


# ── MAIN ────────────────────────────────────────────────────────────
if __name__ == '__main__':
    create_cloud_workbook()
    create_emerging_tech_workbook()
    create_digital_services_workbook()
    create_dcra_workbook()
    create_nra_workbook()
    print('\nAll 5 IESP AWP workbooks generated (v4 — domain-level conclusions).')
